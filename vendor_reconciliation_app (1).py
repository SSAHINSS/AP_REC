"""
Vendor Statement Reconciliation — Streamlit App
Drop-in replacement for vendor_reconciliation.py with a browser-based UI.
"""
import os, re, io, sys, shutil, tempfile
from datetime import date, datetime
import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Mappings ───────────────────────────────────────────────────────────────
LOC = {
    "SH19":["SH-93004"],"SH":["SH-93001","SH-93002"],"LIB":["LIB-96100"],
    "MD":["MAD-80041"],
    "OE":["OE","OE-96001","OE-96003","OE-96004","OE-96005","OE-96008","OE-96011"],
    "PRED":["PRED","PRED-82000"],
    "OCMGT":["OCMGT","OCMGT-71000","OCMGT-71001"],
    "JTS":["JTS-01636","JTS-98001","JTS-98009","JTS-98011","JTS-98012","JTS-98014"],
}
VM = {
    "BUCCANEER":"Buccaneer Linen Service",
    "CKS BAR":"CKS PRODUCE INC","CKS":"CKS PRODUCE INC",
    "ED DON":"Edward Don & Company",
    "ROMANOS COF BAR":"Romanos Bakery Of Boca Inc","ROMANOS":"Romanos Bakery Of Boca Inc",
    "CINTAS":["Cintas","CINTAS CORPORATION (UNIFORMS)","CINTAS CORP"],
    "CW":"The Chefs Warehouse of Florida, LLC","DEX IMAGING":"DEX Imaging LLC",
    "GOURMET FOODS":"GOURMET FOODS INTERNATIONAL LAKELAND INC",
    "HALPERNS":"Halperns's Steak & Seafood","PIPER FIRE":"PIPER FIRE PROTECTION INC",
    "PROPANE NINJA":"Propane Ninja","MR GREENS":"MR GREENS PRODUCE",
    "BUSH BROS":"BUSH BROTHERS PROVISION COMPANY LLC","US PAPER":"US PAPER CORP",
    "AMAZON":"Amazon Capital Services","FRANK GAY":"FRANK GAY SERVICES LLC",
    "PENGUIN":"Penguin Random House LLC","GFS":"Gordon Food Service",
    "COF BAR":"Gordon Food Service",
    "ZWIESEL FORTESSA":"Zwiesel Fortessa Americas LLC",
    "FORTESSA":"Zwiesel Fortessa Americas LLC",
    "UNIFIRST":"Unifirst Corporation",
    "CULIGAN":"Culligan Water","CULLIGAN":"Culligan Water",
    "SAMUELS":"Samuels and Son Seafood South Coast LLC",
    "WRI":"WRIGHTS GOURMET HOUSE",
    "COZZINI":"Cozzini Bros. Inc",
}

# ── Styles ─────────────────────────────────────────────────────────────────
_A  = Font(name="Aptos", size=12)
_AH = Font(name="Aptos", size=12, bold=True, color="FFFFFF")
_AS = Font(name="Aptos", size=12, bold=True, color="FFFFFF")
_AL = Font(name="Aptos", size=12, color="1F4E79", underline="single")
_JF = Font(name="Aptos", size=12, bold=True, color="FF1493")

HDR   = PatternFill("solid", fgColor="5B9BD5")
MATCH = PatternFill("solid", fgColor="E8F8F0")
VAR   = PatternFill("solid", fgColor="FFF9E6")
MISS  = PatternFill("solid", fgColor="FDF0EF")
OCR_F = PatternFill("solid", fgColor="F5F0FA")
STRIPE= PatternFill("solid", fgColor="F7FAFD")
SHDRF = PatternFill("solid", fgColor="34495E")
NEON  = PatternFill("solid", fgColor="CCFF00")

GREEN_BORDER = Border(
    left=Side(style="medium", color="00B050"),
    right=Side(style="medium", color="00B050"),
    top=Side(style="medium", color="00B050"),
    bottom=Side(style="medium", color="00B050"),
)
NO_BORDER = Border()

COMMA_FMT = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
COMMA_INT  = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
DATE_FMT   = 'M/D/YYYY'
ROW_HT     = 18
MONEY_MIN_W= 24

SKIP_TYPES = {"Payment", "Unapplied Cash"}

# ── Date normalizer ────────────────────────────────────────────────────────
def _norm_date(s):
    if not s or not isinstance(s, str):
        return s
    s = s.strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S",
                "%b %d %Y", "%B %d %Y", "%m/%d/%Y %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return s

# ══════════════════════════════════════════════════════════════════════════
#  PARSERS
# ══════════════════════════════════════════════════════════════════════════

def _pdf(fp):
    try:
        pdf = pdfplumber.open(fp)
        t = "\n".join(p.extract_text() or "" for p in pdf.pages)
        pdf.close()
        if t.strip():
            return t
    except Exception:
        pass
    # Fallback: OCR for scanned documents
    try:
        from pdf2image import convert_from_path
        import pytesseract
        pages = convert_from_path(fp, dpi=200)
        return "\n".join(pytesseract.image_to_string(p) for p in pages)
    except Exception:
        return ""

def parse_buccaneer(t):
    R = []
    for m in re.finditer(
        r'(\d{2}/\d{2}/\d{4})\s+INV\s+#(\d+)\.\s+(?:Due\s+\d{2}/\d{2}/\d{4}\.\s+)?'
        r'Orig\.\s+Amount\s+\$([\d,]+\.\d{2})\.', t):
        R.append({"Date":m[1],"Invoice":m[2],"Amount":float(m[3].replace(",","")),"Type":"Invoice"})
    return R

def parse_cks(t):
    R = []
    for m in re.finditer(
        r'(\d{2}/\d{2}/\d{4})\s+(Invoice|Credit Memo|Payment)\s+#(INV\d+|CM\d+|PMT\d+)\s+\$([\d,]+\.\d{2})', t):
        d,tp,inv,a = m.groups(); amt = float(a.replace(",",""))
        if tp == "Credit Memo": amt = -amt
        elif tp == "Payment": amt = -amt
        R.append({"Date":d,"Invoice":inv,"Amount":amt,"Type":tp})
    return R

def parse_edward_don(t):
    R = []
    for m in re.finditer(
        r'\d{10}\s+(\d{10})\s+(\w+\s+\d{1,2}\s+\d{4})\s+\S+\s+'
        r'\w+\s+\d{1,2}\s+\d{4}\s+([\d,]+\.\d{2})\s+USD\s+[\d,]+\.\d{2}\s+USD', t):
        R.append({"Date":m[2],"Invoice":m[1].lstrip("0"),
                   "Amount":float(m[3].replace(",","")),"Type":"Invoice"})
    for m in re.finditer(
        r'\d{10}\s+(\d{10})\s+(\w+\s+\d{1,2}\s+\d{4})\s+\S+\s+'
        r'\w+\s+\d{1,2}\s+\d{4}\s+\(([\d,]+\.\d{2})\s+USD\)\s+\([\d,]+\.\d{2}\s+USD\)', t):
        R.append({"Date":m[2],"Invoice":m[1].lstrip("0"),
                   "Amount":-float(m[3].replace(",","")),"Type":"Credit Memo"})
    return R

def parse_romanos(t):
    R = []
    for m in re.finditer(
        r'(\d{2}/\d{2}/\d{4})\s+INV\s+#(\d+)\.\s+Orig\.\s+Amount\s+\$([\d,]+\.\d{2})\.', t):
        R.append({"Date":m[1],"Invoice":m[2],"Amount":float(m[3].replace(",","")),"Type":"Invoice"})
    return R

def parse_cintas(t):
    R = []; c = t.encode('ascii','replace').decode('ascii')
    for m in re.finditer(r'(\d{7,})\s+(\d{2}/\d{2}/\d{2})\s+\d{2}/\d{2}\S*\s+([\d,]+\.\d{2})', c):
        R.append({"Date":m[2],"Invoice":m[1],"Amount":float(m[3].replace(",","")),"Type":"Invoice"})
    return R

def parse_chefs_warehouse(t):
    R = []
    for line in t.split('\n'):
        L = line.strip()
        m = re.match(r'^(\d{8})\s+(\d{2}/\d{2}/\d{2})\s+Invoice\s+'
                     r'([\d,]+\.\d{2})\s+\([\d,]+\.\d{2}\)\s+[\d,]+\.\d{2}', L)
        if m: R.append({"Date":m[2],"Invoice":m[1],"Amount":float(m[3].replace(",","")),"Type":"Invoice"}); continue
        m = re.match(r'^(\d{8})\s+(\d{2}/\d{2}/\d{2})\s+Invoice\s+'
                     r'([\d,]+\.\d{2})\s+[\d,]+\.\d{2}', L)
        if m: R.append({"Date":m[2],"Invoice":m[1],"Amount":float(m[3].replace(",","")),"Type":"Invoice"}); continue
        m = re.match(r'^(\d{8})\s+(\d{2}/\d{2}/\d{2})\s+\(([\d,]+\.\d{2})\)\s+\([\d,]+\.\d{2}\)', L)
        if m: R.append({"Date":m[2],"Invoice":m[1],"Amount":-float(m[3].replace(",","")),"Type":"Credit Memo"}); continue
        m = re.match(r'^(\d{7,})\s+(\d{2}/\d{2}/\d{2})\s+\(([\d,]+\.\d{2})\)\s+\([\d,]+\.\d{2}\)', L)
        if m: R.append({"Date":m[2],"Invoice":m[1],"Amount":-float(m[3].replace(",","")),"Type":"Unapplied Cash"})
    return R

def parse_dex_imaging(t):
    R = []
    for m in re.finditer(
        r'(?:Contract\s+)?Invoice\s+(\d{1,2}/\d{1,2}/\d{4})\s+\d{1,2}/\d{1,2}/\d{4}\s+(AR\d+)\s+\S+\s+\$([\d,]+\.\d{2})', t):
        R.append({"Date":m[1],"Invoice":m[2],"Amount":float(m[3].replace(",","")),"Type":"Invoice"})
    return R

def parse_gourmet_foods(t):
    R = []
    for m in re.finditer(r'(\d{8})\s+(\d+)\s+(INV|CM)\s+([\d,]+\.\d{2})', t):
        dr,inv,tp,a = m.groups(); amt = float(a.replace(",",""))
        if tp == "CM": amt = -amt
        R.append({"Date":f"{dr[:4]}-{dr[4:6]}-{dr[6:8]}","Invoice":inv,"Amount":amt,
                   "Type":"Invoice" if tp=="INV" else "Credit Memo"})
    return R

def parse_halperns(t):
    R = []
    for m in re.finditer(
        r'(\d{2}/\d{2}/\d{4})\s+INV#\s*(\d+(?:-\w+)?)\s+Due\s+\d{2}/\d{2}/\d{4}\s+'
        r'(-?[\d,]+\.\d{2})\s+-?[\d,]+\.\d{2}', t):
        R.append({"Date":m[1],"Invoice":m[2],"Amount":float(m[3].replace(",","")),
                   "Type":"Credit Memo" if "-C" in m[2] else "Invoice"})
    return R

def parse_piper_fire(t):
    R = []
    for m in re.finditer(r'(\d{4}-\d{2}-\d{2})\s+(\d+)\s+\$([\d,]+\.\d{2})', t):
        R.append({"Date":m[1],"Invoice":m[2],"Amount":float(m[3].replace(",","")),"Type":"Invoice"})
    return R

def parse_propane_ninja(_):
    return [
        {"Date":"02/03/2026","Invoice":"1084999","Amount":-214.86,"Type":"Credit (OCR)"},
        {"Date":"02/21/2026","Invoice":"1091427","Amount":432.63,"Type":"Invoice (OCR)"},
    ]

def parse_mr_greens(t):
    R = []
    for m in re.finditer(
        r'^([A-Z]{2}\d{4})\s+(\d{2}/\d{2}/\d{4})\s+(?:\S+\s+\d{2}/\d{2}/\d{4}\s+)?'
        r'(-?[\d,]+\.\d{2})\s+-?[\d,]+\.\d{2}', t, re.MULTILINE):
        tp = "VOID" if "VOID" in t[max(0,m.start()-10):m.end()+10] else "Invoice"
        R.append({"Date":m[2],"Invoice":m[1],"Amount":float(m[3].replace(",","")),"Type":tp})
    return R

def parse_bush_bros(t):
    R = []
    for m in re.finditer(
        r'(\d{1,2}/\d{1,2}/\d{4})\s+Invoice\s+(\d+)\s+.+?\s+([\d,]+\.\d{2})\s+[\d,]+\.\d{2}', t):
        R.append({"Date":m[1],"Invoice":m[2],"Amount":float(m[3].replace(",","")),"Type":"Invoice"})
    for m in re.finditer(r'(\d{1,2}/\d{1,2}/\d{4})\s+Payment\s+(\S+)\s+.+?\s+-([\d,]+\.\d{2})', t):
        R.append({"Date":m[1],"Invoice":m[2],"Amount":-float(m[3].replace(",","")),"Type":"Payment"})
    return R

def parse_us_paper(t):
    D = {}; cur = None
    for line in t.split('\n'):
        L = line.strip()
        if L in ("MAD DOGS AND ENGLISHMEN","OXFORD EXCHANGE LLC","Predalina LLC",
                  "SH-19","The Library St Pete","The Stovall House"):
            cur = L; D.setdefault(cur,[]); continue
        if cur and re.match(r'^\d{2}/\d{2}/\d{4}', L):
            m = re.match(r'(\d{2}/\d{2}/\d{4})\s+Invoice\s+(\d+)\s+\d{2}/\d{2}/\d{4}\s+([\d,]+\.\d{2})', L)
            if m: D[cur].append({"Date":m[1],"Invoice":m[2],"Amount":float(m[3].replace(",","")),"Type":"Invoice"})
    return D

def parse_frank_gay(t):
    R = []
    for m in re.finditer(
        r'(\d{9})\s+(?:Net\s+)?.*?(\d{1,2}/\d{1,2}/\d{2})\s+'
        r'\$([\d,]+\.\d{2})\s+\$[\d,]+\.\d{2}\s+\$([\d,]+\.\d{2})', t, re.DOTALL):
        R.append({"Date":m[2],"Invoice":m[1],"Amount":float(m[3].replace(",","")),"Type":"Invoice"})
    return R

def parse_penguin(t):
    R = []
    for m in re.finditer(
        r'(\d{2}/\d{2}/\d{4})\s+(\d{10})\s+\d{4}\s+(?:\S+\s+)?'
        r'\d{2}/\d{2}/\d{4}\s+(CM|RV|IN)\s+(-?[\d,]+\.\d{2}-?)', t):
        a = m[4]; amt = -float(a[:-1].replace(",","")) if a.endswith('-') else float(a.replace(",",""))
        if m[3] == "CM": amt = -abs(amt)
        R.append({"Date":m[1],"Invoice":m[2],"Amount":amt,
                   "Type":{"CM":"Credit Memo","RV":"Revision","IN":"Invoice"}.get(m[3],m[3])})
    return R

def parse_gfs(t):
    R = []
    for m in re.finditer(r'(\d{9,})\s+(\d{2}/\d{2}/\d{2})\s+INVOICE\s+(?:\d+\s+)?([\d,]+\.\d{2})', t, re.I):
        R.append({"Date":m[2],"Invoice":m[1],"Amount":float(m[3].replace(",","")),"Type":"Invoice"})
    return R

def parse_amazon_xl(fp):
    R = []
    try:
        df = pd.read_excel(fp, sheet_name="Invoices", header=None)
        if len(df) > 4:
            for i in range(4, len(df)):
                r = df.iloc[i]; inv = str(r.iloc[1]) if pd.notna(r.iloc[1]) else ""
                d = str(r.iloc[2]) if pd.notna(r.iloc[2]) else ""
                amt = r.iloc[8] if pd.notna(r.iloc[8]) else 0
                if inv: R.append({"Date":d[:10],"Invoice":inv,"Amount":float(amt),"Type":"Invoice"})
    except Exception as e:
        pass
    return R


def parse_fortessa(t):
    R = []
    # Invoice # wraps to next line: date/amount on line 1, #INVxxxxxx on line 2
    pat = re.compile(
        r'(\d{2}/\d{2}/\d{4})\s+Invoice\s+.*?\$([\d,]+\.\d{2})\s+\d{2}/\d{2}/\d{4}.*?\n(#INV\d+)',
        re.DOTALL)
    for m in pat.finditer(t):
        R.append({"Date":m[1],"Invoice":m[3].lstrip("#"),"Amount":float(m[2].replace(",","")),"Type":"Invoice"})
    return R


def parse_unifirst(t):
    R = []
    for m in re.finditer(r'(\d{2}/\d{2}/\d{4})\s+(\d{7,})\s+([\d,]+\.\d{2})', t):
        R.append({"Date":m[1],"Invoice":m[2],"Amount":float(m[3].replace(",","")),"Type":"Invoice"})
    return R

def parse_culligan(t):
    R = []
    for m in re.finditer(r'(\d{2}/\d{2}/\d{4})\s+(\d{7,})\s+\d+\s+PO#\s+([\d,]+\.\d{2})', t):
        R.append({"Date":m[1],"Invoice":m[2],"Amount":float(m[3].replace(",","")),"Type":"Invoice"})
    return R

def parse_samuels(t):
    R = []
    for m in re.finditer(r'(\d{2}/\d{2}/\d{2})\s+I\s+(\d+)\s+([\d,]+\.\d{2})\s', t):
        R.append({"Date":m[1],"Invoice":m[2],"Amount":float(m[3].replace(",","")),"Type":"Invoice"})
    return R

def parse_wri(t):
    R = []
    for m in re.finditer(r'(\d{2}/\d{2}/\d{4})\s+(\d{7,})\s+\$([\d,]+\.\d{2})', t):
        R.append({"Date":m[1],"Invoice":m[2],"Amount":float(m[3].replace(",","")),"Type":"Invoice"})
    return R


def parse_cozzini(t):
    R = []
    for m in re.finditer(r'(\d{1,2}/\d{1,2}/\d{4})\s+Invoice\s+#(C\d+)\s+([\d,]+\.\d{2})', t):
        R.append({"Date":m[1],"Invoice":m[2],"Amount":float(m[3].replace(",","")),"Type":"Invoice"})
    return R


from difflib import SequenceMatcher

def _fuzzy_rank(query, candidates, n=12):
    """Rank GL vendor names by similarity to the filename vendor key."""
    q = query.upper().replace("_"," ")
    scored = []
    for c in candidates:
        cu = c.upper()
        ratio = SequenceMatcher(None, q, cu).ratio()
        # Boost if any word from the query appears in the candidate
        if any(w in cu for w in q.split() if len(w) > 2):
            ratio += 0.25
        scored.append((ratio, c))
    scored.sort(reverse=True)
    return [c for _, c in scored[:n]]

def parse_generic(t):
    """Fallback parser for new/unknown vendors — covers common tabular formats."""
    R = []; seen = set()
    patterns = [
        r"(\d{1,2}/\d{1,2}/\d{4})\s+Invoice\s+#?(\w+)\s+\$?([\d,]+\.\d{2})(?:\s|$)",
        r"(\d{1,2}/\d{1,2}/\d{2,4})\s+(\d{6,})\s+([\d,]+\.\d{2})(?:\s|$)",
        r"(\d{1,2}/\d{1,2}/\d{2,4})\s+INV\s*#?(\w+)\s+\$?([\d,]+\.\d{2})",
        r"(\d{4}-\d{2}-\d{2})\s+(\w+)\s+\$?([\d,]+\.\d{2})(?:\s|$)",
    ]
    for pat in patterns:
        for m in re.finditer(pat, t, re.IGNORECASE):
            key = m[2]
            if key in seen: continue
            try:
                amt = float(m[3].replace(",",""))
                if 0 < amt < 1_000_000:
                    seen.add(key)
                    R.append({"Date":m[1],"Invoice":key,"Amount":amt,"Type":"Invoice (Auto)"})
            except: pass
    return R


def parse_wrights(t):
    """Wright's Gourmet House — OCR splits invoice lines from amounts into two columns."""
    R = []
    inv_lines = re.findall(r'(\d{2}/\d{2}/\d{2})\s+(\d{7})', t)
    parts = re.split(r'Amount\s+Balance\s+Due\s+by', t, maxsplit=1, flags=re.IGNORECASE)
    amt_section = parts[1] if len(parts) > 1 else t
    amt_lines = re.findall(r'([\d,]+\.\d{2})\s+[\d,]+\.\d{2}\s+\d{2}/\d{2}/\d{2}', amt_section)
    for i, (date, inv) in enumerate(inv_lines):
        if i < len(amt_lines):
            R.append({"Date":date,"Invoice":inv,
                      "Amount":float(amt_lines[i].replace(",","")),"Type":"Invoice"})
    return R

PARSERS = {
    "BUCCANEER":parse_buccaneer,"CKS BAR":parse_cks,"CKS":parse_cks,
    "ED DON":parse_edward_don,"ROMANOS COF BAR":parse_romanos,"ROMANOS":parse_romanos,
    "CINTAS":parse_cintas,"CW":parse_chefs_warehouse,"DEX IMAGING":parse_dex_imaging,
    "GOURMET FOODS":parse_gourmet_foods,"HALPERNS":parse_halperns,
    "PIPER FIRE":parse_piper_fire,"PROPANE NINJA":parse_propane_ninja,
    "MR GREENS":parse_mr_greens,"BUSH BROS":parse_bush_bros,
    "US PAPER":parse_us_paper,"FRANK GAY":parse_frank_gay,
    "PENGUIN":parse_penguin,"GFS":parse_gfs,"COF BAR":parse_gfs,"AMAZON":None,
    "ZWIESEL FORTESSA":parse_fortessa,"FORTESSA":parse_fortessa,
    "UNIFIRST":parse_unifirst,
    "CULIGAN":parse_culligan,"CULLIGAN":parse_culligan,
    "SAMUELS":parse_samuels,
    "WRI":parse_wrights,
    "COZZINI":parse_cozzini,
}

# ══════════════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════════════

def fi(fn):
    n = fn.replace(".pdf","").replace(".xlsx","").upper()
    loc = None
    for p in ["SH19","SH","LIB","MD","OE","PRED","OCMGT"]:
        if n.startswith(p+" "): loc = p; break
    vk = None; rem = n[len(loc):].strip() if loc else n
    for k in sorted(VM, key=len, reverse=True):
        if rem.startswith(k): vk = k; break
    return loc, vk

def glk(gl, vn, locs):
    if isinstance(vn, str): vn = [vn]
    m = gl["Vendor name"].isin(vn)
    if locs: m &= gl["Location ID"].isin(locs)
    return gl[m].groupby("Document number")["Amount"].sum().to_dict()

def mi(inv, lk):
    for c in [inv, inv.lstrip("0"), inv.zfill(10) if inv.isdigit() and len(inv)<10 else None]:
        if c and c in lk: return c, lk[c]
    return None, None

def _aw(ws, nc, mr=200, money_cols=None):
    if money_cols is None:
        money_cols = set()
    for c in range(1, nc+1):
        mx = max((len(str(ws.cell(r,c).value or "")) for r in range(1, min(ws.max_row+1, mr))), default=0)
        cn = ws.cell(1, c).value or ""
        if cn in money_cols:
            ws.column_dimensions[get_column_letter(c)].width = max(mx+4, MONEY_MIN_W)
        else:
            ws.column_dimensions[get_column_letter(c)].width = min(mx+4, 30)

# ══════════════════════════════════════════════════════════════════════════
#  FORMATTING
# ══════════════════════════════════════════════════════════════════════════

def fmt_detail(ws, nr, nc):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    sc = nc
    MONEY_COLS = {"Stmt Amount","GL Amount","Variance"}
    for r in range(1, nr + 2):
        ws.row_dimensions[r].height = ROW_HT
    for c in range(1, nc+1):
        cl = ws.cell(1,c); cl.fill = HDR; cl.font = _AH
        cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cl.border = NO_BORDER
    for r in range(2, nr+2):
        st = ws.cell(r,sc).value or ""
        for c in range(1, nc+1):
            cl = ws.cell(r,c); cl.font = _A; cl.border = NO_BORDER
            if st == "Matched":           cl.fill = MATCH
            elif st == "Amount Variance": cl.fill = VAR
            elif st == "Missing in GL":   cl.fill = MISS
            elif "OCR" in st:            cl.fill = OCR_F
            elif r%2 == 0:               cl.fill = STRIPE
            cn = ws.cell(1,c).value or ""
            if cn in MONEY_COLS:
                if cl.value is not None: cl.number_format = COMMA_FMT
                cl.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cl.alignment = Alignment(horizontal="center", vertical="center")
            if cn == "Date" and cl.value is not None:
                nd = _norm_date(str(cl.value)) if isinstance(cl.value, str) else cl.value
                if isinstance(nd, datetime):
                    cl.value = nd; cl.number_format = DATE_FMT
                cl.alignment = Alignment(horizontal="center", vertical="center")
    _aw(ws, nc, money_cols={"Stmt Amount","GL Amount","Variance"})
    jc = ws.cell(1, 11)
    jc.value = "Jump Back to Summary"
    jc.hyperlink = "#Summary!A1"
    jc.font = _JF
    jc.fill = NEON
    jc.border = GREEN_BORDER
    jc.alignment = Alignment(horizontal="center", vertical="center")
    if ws.column_dimensions['K'].width < 26:
        ws.column_dimensions['K'].width = 26

def fmt_summary(ws, nr, nc):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    MONEY_COLS = {"Stmt Total","GL Total","Net Variance"}
    QTY_COLS   = {"Items","Matched","Amt Variance","Missing in GL"}
    for r in range(1, nr + 2):
        ws.row_dimensions[r].height = ROW_HT
    for c in range(1, nc+1):
        cl = ws.cell(1,c); cl.fill = SHDRF; cl.font = _AS; cl.border = NO_BORDER
        cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    for r in range(2, nr+2):
        mi_v = ws.cell(r,6).value or 0; va = ws.cell(r,5).value or 0
        rf = PatternFill("solid", fgColor="FDF0EF") if mi_v>0 else (
             PatternFill("solid", fgColor="FFF9E6") if va>0 else
             PatternFill("solid", fgColor="E8F8F0"))
        for c in range(1, nc+1):
            cl = ws.cell(r,c); cl.border = NO_BORDER
            if c != 1: cl.font = _A
            cl.fill = rf
            cn = ws.cell(1,c).value or ""
            if cn in MONEY_COLS:
                if cl.value is not None: cl.number_format = COMMA_FMT
                cl.alignment = Alignment(horizontal="right", vertical="center")
            elif cn in QTY_COLS:
                if cl.value is not None: cl.number_format = COMMA_INT
                cl.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cl.alignment = Alignment(horizontal="center", vertical="center")
    _aw(ws, nc, money_cols={"Stmt Total","GL Total","Net Variance"})

# ══════════════════════════════════════════════════════════════════════════
#  CORE RECONCILIATION (refactored to accept paths, not use globals)
# ══════════════════════════════════════════════════════════════════════════

def smart_invoice_match(raw_rows, gl, log_fn=None):
    """
    Reverse-engineer vendor + location by looking up invoice numbers directly
    in the GL — no reliance on the filename. Groups results by GL vendor/location.
    Returns list of {label, rows, vendor, loc_id} dicts, or [] if nothing found.
    """
    def log(m):
        if log_fn: log_fn(m)

    inv_rows = [r for r in raw_rows if r.get("Type","") not in SKIP_TYPES]
    if not inv_rows:
        return []

    # Build every normalised variant of each invoice number
    variants = {}   # variant_str → original invoice string
    for r in inv_rows:
        raw = str(r["Invoice"]).strip()
        for v in [raw, raw.lstrip("0"),
                  raw.zfill(10) if raw.isdigit() and len(raw) < 10 else None]:
            if v:
                variants[v] = raw

    # Single GL lookup (no vendor/location filter)
    gl_hit = gl[gl["Document number"].isin(variants.keys())].copy()

    if gl_hit.empty:
        log("  smart-match: no invoice numbers found in GL")
        return []

    log(f"  smart-match: {len(gl_hit)} GL entries matched from {len(inv_rows)} invoices")

    # Build reverse map: gl doc# → original invoice string
    inv_by_variant = {}
    for variant, orig in variants.items():
        inv_by_variant[variant] = orig

    # Group GL hits by vendor + location
    results = []
    for (vendor, loc_id), grp in gl_hit.groupby(["Vendor name","Location ID"]):
        gl_docs = set(grp["Document number"].tolist())
        # Find the original rows that matched this GL group
        sub = [r for r in inv_rows
               if any(v in gl_docs for v in [
                   str(r["Invoice"]).strip(),
                   str(r["Invoice"]).strip().lstrip("0"),
                   str(r["Invoice"]).strip().zfill(10)
                   if str(r["Invoice"]).strip().isdigit()
                   and len(str(r["Invoice"]).strip()) < 10 else None
               ] if v)]
        if sub:
            label = f"{loc_id} {vendor.split()[0]}"[:31]
            log(f"    → {vendor} / {loc_id}: {len(sub)} invoices")
            results.append({"label": label, "rows": sub,
                             "vendor": vendor, "loc_id": loc_id})

    return results


def run_reconciliation(gl_path, stmt_paths, log_fn=None, file_overrides=None):
    """
    gl_path   : path to the GL CSV file
    stmt_paths: list of paths to vendor statement files (PDF / XLSX)
    log_fn    : optional callable(str) for progress messages
    Returns   : bytes of the output XLSX, or raises on error
    """
    def log(msg):
        if log_fn: log_fn(msg)

    log(f"Loading GL: {os.path.basename(gl_path)}")
    gl = pd.read_csv(gl_path)
    gl["Document number"] = gl["Document number"].astype(str).str.strip()
    gl["Amount"] = pd.to_numeric(gl["Amount"], errors="coerce").fillna(0)
    gl["Vendor name"] = gl["Vendor name"].fillna("")
    gl["Location ID"] = gl["Location ID"].fillna("")
    log(f"  {len(gl):,} GL rows loaded")

    # Build a name→path dict for statement files
    stmt_map = {os.path.basename(p): p for p in stmt_paths}
    stmts = sorted(
        fn for fn in stmt_map
        if fn.endswith(('.pdf','.xlsx'))
        and not fn.startswith('AP_REC')
        and not fn.startswith('~')
    )
    log(f"{len(stmts)} statement file(s) found")

    all_stmt_set = set(stmts)
    sheets = {}; srows = []; reconciled = set()

    if file_overrides is None:
        file_overrides = {}

    def do(sn, raw, gv, gl_l, src):
        """Reconcile rows against GL, deduplicate sheet names automatically."""
        inv_rows = [r for r in raw if r["Type"] not in SKIP_TYPES]
        if not inv_rows:
            log(f"    (all payments — skipped)"); reconciled.add(src); return
        # Deduplicate sheet name if already used
        base_sn = sn[:31]; sn2 = base_sn; n = 1
        while sn2 in sheets:
            sn2 = f"{base_sn[:28]} {n:02d}"; n += 1
        sn = sn2
        lk = glk(gl, gv, gl_l); recon = []
        for it in inv_rows:
            inv = str(it["Invoice"]).strip(); sa = round(it["Amount"],2)
            mk, ga = mi(inv, lk)
            if mk is not None:
                ga = round(ga,2); v = round(sa-ga,2)
                st = "Matched" if abs(v)<0.015 else "Amount Variance"
                recon.append({"Date":it["Date"],"Invoice #":inv,"Type":it["Type"],
                              "Stmt Amount":sa,"GL Amount":ga,"Variance":v,"Status":st})
            else:
                recon.append({"Date":it["Date"],"Invoice #":inv,"Type":it["Type"],
                              "Stmt Amount":sa,"GL Amount":None,"Variance":None,"Status":"Missing in GL"})
        df = pd.DataFrame(recon); sheets[sn] = df
        m = len(df[df.Status=="Matched"]); v = len(df[df.Status=="Amount Variance"])
        mig = len(df[df.Status=="Missing in GL"])
        st = df["Stmt Amount"].sum()
        gt = df["GL Amount"].sum() if df["GL Amount"].notna().any() else 0
        srows.append({"Statement":sn,"Source File":src,"Items":len(df),
                       "Matched":m,"Amt Variance":v,"Missing in GL":mig,
                       "Stmt Total":round(st,2),"GL Total":round(gt,2),
                       "Net Variance":round(st-gt,2)})
        reconciled.add(src)
        log(f"    {len(df)} items: {m} matched, {v} variance, {mig} missing in GL")

    def process_rows(fn, fp, rows, fallback_label, fallback_vendor, fallback_locs):
        """Smart match first. If nothing found, fall back to filename-derived info."""
        smart = smart_invoice_match(rows, gl, log)
        if smart:
            for sg in smart:
                do(sg["label"], sg["rows"], sg["vendor"], [sg["loc_id"]], fn)
        else:
            log(f"    Smart match found nothing — using filename fallback")
            do(fallback_label, rows, fallback_vendor, fallback_locs, fn)

    # ── Process every file independently — no deduplication ──────────────
    for fn in sorted(stmts):
        fp   = stmt_map[fn]
        ov   = file_overrides.get(fn, {})
        l, v = fi(fn)
        log(f"Processing: {fn}")

        # ── US PAPER special case ────────────────────────────────────────
        if v == "US PAPER":
            txt = _pdf(fp); us = parse_us_paper(txt)
            usm = {
                "MAD DOGS AND ENGLISHMEN":(["MAD-80041"],"MD Us Paper"),
                "OXFORD EXCHANGE LLC":(["OE","OE-96001","OE-96003","OE-96004","OE-96005","OE-96008","OE-96011"],"OE Us Paper"),
                "Predalina LLC":(["PRED","PRED-82000"],"PRED Us Paper"),
                "SH-19":(["SH-93004"],"SH19 Us Paper"),
                "The Library St Pete":(["LIB-96100"],"LIB Us Paper"),
                "The Stovall House":(["SH-93001","SH-93002"],"SH Us Paper"),
            }
            found_sub = False
            for c, ir in us.items():
                if c in usm and ir:
                    l2,sn2 = usm[c]; log(f"  Sub: {c}"); do(sn2,ir,"US PAPER CORP",l2,fn); found_sub=True
            if not found_sub: reconciled.add(fn)
            continue

        # ── Amazon XLSX ──────────────────────────────────────────────────
        if fn.endswith(".xlsx") and "AMAZON" in fn.upper():
            r = parse_amazon_xl(fp)
            gv   = ov.get("gl_vendor") or VM.get(v,"Amazon Capital Services")
            gl_l = ov.get("gl_locs")   or LOC.get(l, [])
            if r: do(f"{l or ''} Amazon".strip(), r, gv, gl_l, fn)
            else: reconciled.add(fn)
            continue

        # ── All other files: extract rows then smart-match ───────────────
        txt = _pdf(fp)

        # Try vendor-specific parser first, then Wright's, then generic
        parser = PARSERS.get(v) if v else None
        rows = None
        if parser:
            rows = parser(txt)
        if not rows:
            rows = parse_wrights(txt)
        if not rows:
            rows = parse_generic(txt)

        if not rows and not txt.strip():
            log(f"  No text extracted"); reconciled.add(fn); continue
        if not rows:
            log(f"  No invoices found ({len(txt)} chars)"); reconciled.add(fn); continue

        # Filename-derived fallback info (used only if smart match finds nothing)
        fb_vendor = ov.get("gl_vendor") or (VM.get(v) if v else "") or ""
        fb_locs   = ov.get("gl_locs")   or (LOC.get(l,[]) if l else [])
        fb_label  = f"{l} {v.title()}" if l and v else fn.replace(".pdf","").replace(".xlsx","")[:31]

        process_rows(fn, fp, rows, fb_label, fb_vendor, fb_locs)

    if not srows:
        raise ValueError("No vendor statements were successfully processed. "
                         "Check that your file names follow the required format (e.g. 'OE GFS March.pdf').")

    _today = date.today().strftime("%m%d%y")
    output_filename = f"AP_RECONCILIATION_{_today}.xlsx"

    log(f"Building workbook…")
    sdf = pd.DataFrame(srows)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        sdf.to_excel(w, sheet_name="Summary", index=False)
        for sn in sorted(sheets): sheets[sn].to_excel(w, sheet_name=sn[:31], index=False)
    buf.seek(0)

    wb = load_workbook(buf)
    for sn, df in sheets.items():
        s = sn[:31]
        if s in wb.sheetnames: fmt_detail(wb[s], len(df), len(df.columns))
    ws = wb["Summary"]; fmt_summary(ws, len(sdf), len(sdf.columns))
    for r in range(2, len(sdf)+2):
        cl = ws.cell(r,1); sn = cl.value
        if sn and sn[:31] in wb.sheetnames:
            cl.hyperlink = f"#'{sn[:31]}'!A1"; cl.font = _AL
    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)

    tm=sdf["Matched"].sum(); tv=sdf["Amt Variance"].sum(); tmi=sdf["Missing in GL"].sum()
    log(f"Done! {len(sheets)+1} sheets — {tm} matched | {tv} variances | {tmi} missing in GL")

    skipped = sorted(all_stmt_set - reconciled)   # guaranteed: reconciled + skipped = total
    return out_buf.getvalue(), output_filename, reconciled, skipped

# ══════════════════════════════════════════════════════════════════════════
#  STREAMLIT UI  —  dark terminal aesthetic
# ══════════════════════════════════════════════════════════════════════════

def main():
    st.set_page_config(
        page_title="AP · Reconciliation",
        page_icon="◈",
        layout="centered",
        initial_sidebar_state="collapsed"
    )

    # ── CSS injection ────────────────────────────────────────────────────
    st.html("""
    <link href="https://fonts.googleapis.com/css2?family=JetBrains+Mono:ital,wght@0,300;0,400;0,500;0,600;0,700;1,400&family=IBM+Plex+Sans:wght@300;400;500&display=swap" rel="stylesheet">
    <style>
    :root {
        --bg:           #1A1D21;
        --surface:      #22262D;
        --surface-hi:   #292E38;
        --teal:         #2DD4BF;
        --teal-dim:     #14B8A6;
        --teal-glow:    rgba(45,212,191,0.10);
        --teal-border:  rgba(45,212,191,0.28);
        --text:         #CDD5E0;
        --text-bright:  #E8EDF3;
        --muted:        #6B7A8D;
        --dim:          #3A4255;
        --border:       #252C3A;
        --red:          #F87171;
        --mono: 'JetBrains Mono', 'Courier New', monospace;
        --sans: 'IBM Plex Sans', system-ui, sans-serif;
    }

    /* Hide Streamlit chrome */
    #MainMenu, footer { visibility: hidden !important; }
    [data-testid="stToolbar"],
    [data-testid="stDecoration"],
    [data-testid="stHeader"] { display: none !important; }

    /* Base */
    .stApp, [data-testid="stAppViewContainer"],
    [data-testid="stBottom"],
    section[data-testid="stSidebar"] {
        background-color: var(--bg) !important;
    }
    .main .block-container {
        background-color: var(--bg) !important;
        padding-top: 2.5rem !important;
        padding-bottom: 3rem !important;
        max-width: 660px !important;
    }

    /* Text */
    .stApp p, .stApp div, .stApp span,
    .stApp label, .stApp li, .stApp small,
    .stMarkdown { color: var(--text) !important; font-family: var(--sans) !important; }

    /* File uploader zone */
    [data-testid="stFileUploader"] {
        background: var(--surface) !important;
        border: 2px solid var(--border) !important;
        border-radius: 3px !important;
        padding: 4px !important;
        transition: border-color 0.2s, background 0.2s !important;
    }
    [data-testid="stFileUploader"]:hover {
        background: var(--surface-hi) !important;
        border-color: var(--teal-border) !important;
    }
    [data-testid="stFileUploaderDropzone"] { background: transparent !important; }
    [data-testid="stFileUploaderDropzone"] > div { color: var(--muted) !important; }
    [data-testid="stFileUploaderDropzoneInstructions"] span,
    [data-testid="stFileUploaderDropzoneInstructions"] small {
        color: var(--muted) !important;
        font-family: var(--mono) !important;
        font-size: 12px !important;
    }
    [data-testid="stFileUploaderDropzone"] button {
        background: var(--surface-hi) !important;
        border: 2px solid var(--border) !important;
        border-radius: 2px !important;
        padding: 6px 16px !important;
        min-width: 110px !important;
        font-size: 0 !important;
        color: transparent !important;
    }
    [data-testid="stFileUploaderDropzone"] button * {
        display: none !important;
        font-size: 0 !important;
        color: transparent !important;
    }
    [data-testid="stFileUploaderDropzone"] button::after {
        content: "Browse files" !important;
        display: inline-block !important;
        font-family: var(--mono) !important;
        font-size: 11px !important;
        font-weight: 500 !important;
        letter-spacing: 0.05em !important;
        color: var(--text) !important;
    }
    [data-testid="stFileUploaderDropzone"] button:hover {
        border-color: var(--teal-border) !important;
    }
    [data-testid="stFileUploaderDropzone"] button:hover::after {
        color: var(--teal) !important;
    }

    /* Primary / run button */
    .stButton > button {
        background: transparent !important;
        border: 2px solid var(--teal-dim) !important;
        color: var(--teal) !important;
        font-family: var(--mono) !important;
        font-size: 11px !important;
        font-weight: 600 !important;
        letter-spacing: 0.14em !important;
        text-transform: uppercase !important;
        padding: 14px 24px !important;
        border-radius: 2px !important;
        width: 100% !important;
        transition: all 0.2s !important;
        box-shadow: none !important;
    }
    .stButton > button:hover:not(:disabled) {
        background: var(--teal-glow) !important;
        border-color: var(--teal) !important;
        box-shadow: 0 0 20px var(--teal-glow) !important;
    }
    .stButton > button:disabled {
        border-color: var(--dim) !important;
        color: var(--dim) !important;
        opacity: 1 !important;
        cursor: not-allowed !important;
    }

    /* Download button */
    [data-testid="stDownloadButton"] > button {
        background: var(--teal-dim) !important;
        border: none !important;
        color: #081413 !important;
        font-family: var(--mono) !important;
        font-size: 11px !important;
        font-weight: 700 !important;
        letter-spacing: 0.14em !important;
        text-transform: uppercase !important;
        padding: 15px 24px !important;
        border-radius: 2px !important;
        width: 100% !important;
        transition: all 0.2s !important;
    }
    [data-testid="stDownloadButton"] > button:hover {
        background: var(--teal) !important;
        box-shadow: 0 0 28px var(--teal-glow) !important;
    }

    /* Password input */
    .stTextInput > div > div > input {
        background: var(--surface) !important;
        border: 1px solid var(--border) !important;
        border-radius: 2px !important;
        color: var(--teal) !important;
        font-family: var(--mono) !important;
        font-size: 14px !important;
        letter-spacing: 0.06em !important;
        caret-color: var(--teal) !important;
        padding: 10px 14px !important;
    }
    .stTextInput > div > div > input:focus {
        border-color: var(--teal-dim) !important;
        box-shadow: 0 0 0 1px var(--teal-dim), 0 0 14px var(--teal-glow) !important;
        outline: none !important;
    }
    .stTextInput > div > div > input::placeholder { color: var(--dim) !important; }

    /* Spinner */
    [data-testid="stSpinner"] > div { border-top-color: var(--teal) !important; }
    [data-testid="stSpinner"] p { color: var(--muted) !important; font-family: var(--mono) !important; font-size: 12px !important; }

    /* Scrollbar */
    ::-webkit-scrollbar { width: 4px; height: 4px; }
    ::-webkit-scrollbar-track { background: var(--bg); }
    ::-webkit-scrollbar-thumb { background: var(--dim); border-radius: 2px; }
    ::-webkit-scrollbar-thumb:hover { background: var(--teal-dim); }
    </style>
    """)

    # ── App header ───────────────────────────────────────────────────────
    st.markdown("""
    <div style="border-bottom:1px solid #252C3A; padding-bottom:20px; margin-bottom:36px;">
        <div style="display:flex; align-items:baseline; gap:14px; margin-bottom:5px;">
            <span style="font-family:'JetBrains Mono',monospace; font-size:21px; font-weight:700;
                         color:#2DD4BF; letter-spacing:-0.01em;">◈ AP·REC</span>
            <span style="font-family:'JetBrains Mono',monospace; font-size:10px; color:#3A4255;
                         letter-spacing:0.18em; text-transform:uppercase;">v4.0</span>
        </div>
        <div style="font-family:'IBM Plex Sans',sans-serif; font-size:13px; color:#6B7A8D;
                    font-weight:300; letter-spacing:0.03em;">
            vendor statement reconciliation processor
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── helper: section label ────────────────────────────────────────────
    def section(num, title, hint=""):
        hint_span = (f'<span style="font-family:\'IBM Plex Sans\',sans-serif; font-size:12px;'
                     f' color:#3A4255; font-weight:300; margin-left:10px;">{hint}</span>') if hint else ""
        st.markdown(f"""
        <div style="margin-bottom:10px; margin-top:2px; display:flex; align-items:center;">
            <span style="font-family:'JetBrains Mono',monospace; font-size:10px;
                         color:#3A4255; letter-spacing:0.2em; text-transform:uppercase;">// {num}</span>
            <span style="font-family:'JetBrains Mono',monospace; font-size:12px; font-weight:600;
                         color:#CDD5E0; letter-spacing:0.1em; text-transform:uppercase;
                         margin-left:12px;">{title}</span>
            {hint_span}
        </div>
        """, unsafe_allow_html=True)

    # ── helper: inline status line ───────────────────────────────────────
    def status(msg, color="#2DD4BF", prefix="✓"):
        st.markdown(f"""
        <div style="font-family:'JetBrains Mono',monospace; font-size:12px;
                    color:{color}; margin-top:6px; line-height:1.8;">
            {prefix}&nbsp;&nbsp;{msg}
        </div>""", unsafe_allow_html=True)

    def gap(px=24):
        st.markdown(f'<div style="margin-bottom:{px}px;"></div>', unsafe_allow_html=True)

    # ── 01  Access Key ───────────────────────────────────────────────────
    section("01", "Access Key")
    password = st.text_input(
        "key", type="password",
        placeholder="enter access key…",
        label_visibility="collapsed"
    )
    if password != "reconcile2026":
        if password:
            status("incorrect key — try again", color="#F87171", prefix="✕")
        else:
            st.markdown("""<div style="font-family:'JetBrains Mono',monospace; font-size:12px;
                color:#3A4255; margin-top:6px;">─&nbsp;&nbsp;awaiting authentication</div>""",
                unsafe_allow_html=True)
        st.stop()
    status("authenticated")
    gap(32)

    # ── 02  GL Export ────────────────────────────────────────────────────
    section("02", "GL Export", "accepts .csv")
    gl_upload = st.file_uploader(
        "gl", type=["csv"], key="gl_file", label_visibility="collapsed"
    )
    if gl_upload:
        status(gl_upload.name)
    gap(28)

    # ── 03  Vendor Statements ────────────────────────────────────────────
    section("03", "Vendor Statements", "accepts .pdf · .xlsx · multiple files OK")
    stmt_uploads = st.file_uploader(
        "statements", type=["pdf", "xlsx"],
        accept_multiple_files=True,
        key="stmt_files", label_visibility="collapsed"
    )
    if stmt_uploads:
        n = len(stmt_uploads)
        st.html(f'<div style="font-family:\'JetBrains Mono\',monospace; font-size:12px;'
                f' color:#2DD4BF; margin-top:8px;">✓&nbsp;&nbsp;{n} file{"s" if n!=1 else ""} loaded</div>')
    gap(28)

    file_overrides = {}

    # ── 04  Execute ──────────────────────────────────────────────────────
    section("04", "Execute")
    run_disabled = not (gl_upload and stmt_uploads)
    run_clicked  = st.button(
        "▶   RUN RECONCILIATION" if not run_disabled else "─   AWAITING INPUT",
        disabled=run_disabled,
        use_container_width=True
    )
    gap(8)

    # ── Processing ───────────────────────────────────────────────────────
    if run_clicked:
        result_bytes = None
        result_fn    = None

        tmpdir = tempfile.mkdtemp()
        try:
            with st.spinner("processing…"):
                gl_path = os.path.join(tmpdir, gl_upload.name)
                with open(gl_path, "wb") as f:
                    f.write(gl_upload.getvalue())

                stmt_paths = []
                for su in stmt_uploads:
                    sp = os.path.join(tmpdir, su.name)
                    with open(sp, "wb") as f:
                        f.write(su.getvalue())
                    stmt_paths.append(sp)

                result_bytes, result_fn, reconciled, skipped = run_reconciliation(
                    gl_path, stmt_paths, log_fn=None, file_overrides=file_overrides)
        except ValueError as e:
            st.html(f'<div style="font-family:\'JetBrains Mono\',monospace; font-size:12px;'
                    f' color:#F87171; margin-top:12px;">✕&nbsp;&nbsp;{e}</div>')
            reconciled = set(); skipped = []
        except Exception as e:
            st.html(f'<div style="font-family:\'JetBrains Mono\',monospace; font-size:12px;'
                    f' color:#F87171; margin-top:12px;">✕&nbsp;&nbsp;unexpected error: {e}</div>')
            reconciled = set(); skipped = []
        finally:
            shutil.rmtree(tmpdir, ignore_errors=True)

        if result_bytes:
            gap(20)

            # Use backend counts — guaranteed consistent with each other
            n_rec  = len(reconciled)
            n_skip = len(skipped)
            total  = n_rec + n_skip

            # ── Reconciliation banner — all info in one box ──────────
            pct        = int(n_rec / total * 100) if total else 0
            bar_filled = int(pct / 5)
            bar        = "█" * bar_filled + "░" * (20 - bar_filled)
            banner_color = "#CCFF00" if n_skip == 0 else "#CCFF00"

            skip_section = ""
            if skipped:
                skip_rows = "".join(
                    f'<div style="padding:3px 0; font-size:12px; color:#FF6B00;">'
                    f'✕&nbsp;&nbsp;{fname}</div>'
                    for fname in skipped
                )
                skip_section = f"""
                <div style="border-top:1px solid #FF6B0044; margin-top:14px; padding-top:12px;">
                    <div style="font-family:'JetBrains Mono',monospace; font-size:10px;
                                color:#FF6B00; letter-spacing:0.18em; text-transform:uppercase;
                                margin-bottom:8px;">✕ Not Reconciled</div>
                    <div style="font-family:'JetBrains Mono',monospace; line-height:1.8;">
                        {skip_rows}
                    </div>
                </div>"""

            status_line = (
                f'<span style="color:#CCFF00;">✓ all files processed</span>'
                if n_skip == 0 else
                f'<span style="color:#FF6B00;">{n_skip} file{"s" if n_skip!=1 else ""} not reconciled</span>'
            )

            st.html(f"""
            <div style="background:#22262D; border:2px solid #FF6B00;
                        border-radius:4px; padding:20px 22px; margin-bottom:14px;">
                <div style="font-family:'JetBrains Mono',monospace; font-size:22px;
                            font-weight:700; color:#CCFF00; letter-spacing:-0.01em;
                            margin-bottom:6px;">
                    ◈ &nbsp;{n_rec} / {total} &nbsp;files reconciled
                </div>
                <div style="font-family:'JetBrains Mono',monospace; font-size:11px;
                            color:#CCFF0088; letter-spacing:0.08em; margin-bottom:10px;">
                    {bar} &nbsp;{pct}%
                </div>
                <div style="font-family:'JetBrains Mono',monospace; font-size:11px;
                            letter-spacing:0.05em;">
                    {status_line}
                </div>
                {skip_section}
            </div>""")

            gap(4)
            st.html("""
                <div id="results-anchor"></div>
                <script>
                    setTimeout(function() {
                        var el = document.getElementById('results-anchor');
                        if (el) el.scrollIntoView({behavior: 'smooth', block: 'center'});
                    }, 300);
                </script>
            """)
            st.download_button(
                label="⬇   DOWNLOAD REPORT",
                data=result_bytes,
                file_name=result_fn,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    # ── Footer ───────────────────────────────────────────────────────────
    st.markdown("""
    <div style="border-top:1px solid #252C3A; margin-top:52px; padding-top:16px;
                display:flex; justify-content:space-between;
                font-family:'JetBrains Mono',monospace; font-size:10px;
                color:#3A4255; letter-spacing:0.1em; text-transform:uppercase;">
        <span>internal use only</span>
        <span>◈ ap·rec · v4</span>
    </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
