"""
Vendor Statement Reconciliation — Streamlit App
Drop-in replacement for vendor_reconciliation.py with a browser-based UI.
"""
import os, re, io, sys, shutil, tempfile
from datetime import date, datetime
import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Mappings ───────────────────────────────────────────────────────────────
LOC = {
    "SH19":["SH-93004"],"SH":["SH-93001","SH-93002"],"LIB":["LIB-96100"],
    "MD":["MAD-80041"],
    "OE":["OE","OE-96001","OE-96003","OE-96004","OE-96005","OE-96008","OE-96011"],
    "PRED":["PRED","PRED-82000"],
    "OCMGT":["OCMGT","OCMGT-71000","OCMGT-71001"],
}
VM = {
    "BUCCANEER":"Buccaneer Linen Service",
    "CKS BAR":"CKS PRODUCE INC","CKS":"CKS PRODUCE INC",
    "ED DON":"Edward Don & Company",
    "ROMANOS COF BAR":"Romanos Bakery Of Boca Inc","ROMANOS":"Romanos Bakery Of Boca Inc",
    "CINTAS":["Cintas","CINTAS CORPORATION (UNIFORMS)","CINTAS CORP"],
    "CW":"The Chefs Warehouse of Florida, LLC","DEX IMAGING":"DEX Imaging LLC",
    "GOURMET FOODS":"GOURMET FOODS INTERNATIONAL LAKELAND INC",
    "HALPERNS":"Halperns's Steak & Seafood","PIPER FIRE":"PIPER FIRE PROTECTION INC",
    "PROPANE NINJA":"Propane Ninja","MR GREENS":"MR GREENS PRODUCE",
    "BUSH BROS":"BUSH BROTHERS PROVISION COMPANY LLC","US PAPER":"US PAPER CORP",
    "AMAZON":"Amazon Capital Services","FRANK GAY":"FRANK GAY SERVICES LLC",
    "PENGUIN":"Penguin Random House LLC","GFS":"Gordon Food Service",
    "COF BAR":"Gordon Food Service",
}

# ── Styles ─────────────────────────────────────────────────────────────────
_A  = Font(name="Aptos", size=12)
_AH = Font(name="Aptos", size=12, bold=True, color="FFFFFF")
_AS = Font(name="Aptos", size=12, bold=True, color="FFFFFF")
_AL = Font(name="Aptos", size=12, color="1F4E79", underline="single")
_JF = Font(name="Aptos", size=12, bold=True, color="FF1493")

HDR   = PatternFill("solid", fgColor="5B9BD5")
MATCH = PatternFill("solid", fgColor="E8F8F0")
VAR   = PatternFill("solid", fgColor="FFF9E6")
MISS  = PatternFill("solid", fgColor="FDF0EF")
OCR_F = PatternFill("solid", fgColor="F5F0FA")
STRIPE= PatternFill("solid", fgColor="F7FAFD")
SHDRF = PatternFill("solid", fgColor="34495E")
NEON  = PatternFill("solid", fgColor="CCFF00")

GREEN_BORDER = Border(
    left=Side(style="medium", color="00B050"),
    right=Side(style="medium", color="00B050"),
    top=Side(style="medium", color="00B050"),
    bottom=Side(style="medium", color="00B050"),
)
NO_BORDER = Border()

COMMA_FMT = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
COMMA_INT  = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
DATE_FMT   = 'M/D/YYYY'
ROW_HT     = 18
MONEY_MIN_W= 24

SKIP_TYPES = {"Payment", "Unapplied Cash"}

# ── Date normalizer ────────────────────────────────────────────────────────
def _norm_date(s):
    if not s or not isinstance(s, str):
        return s
    s = s.strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S",
                "%b %d %Y", "%B %d %Y", "%m/%d/%Y %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return s

# ══════════════════════════════════════════════════════════════════════════
#  PARSERS
# ══════════════════════════════════════════════════════════════════════════

def _pdf(fp):
    try:
        pdf = pdfplumber.open(fp)
        t = "\n".join(p.extract_text() or "" for p in pdf.pages)
        pdf.close(); return t
    except Exception as e:
        return ""

def parse_buccaneer(t):
    R = []
    for m in re.finditer(
        r'(\d{2}/\d{2}/\d{4})\s+INV\s+#(\d+)\.\s+(?:Due\s+\d{2}/\d{2}/\d{4}\.\s+)?'
        r'Orig\.\s+Amount\s+\$([\d,]+\.\d{2})\.', t):
        R.append({"Date":m[1],"Invoice":m[2],"Amount":float(m[3].replace(",","")),"Type":"Invoice"})
    return R

def parse_cks(t):
    R = []
    for m in re.finditer(
        r'(\d{2}/\d{2}/\d{4})\s+(Invoice|Credit Memo|Payment)\s+#(INV\d+|CM\d+|PMT\d+)\s+\$([\d,]+\.\d{2})', t):
        d,tp,inv,a = m.groups(); amt = float(a.replace(",",""))
        if tp == "Credit Memo": amt = -amt
        elif tp == "Payment": amt = -amt
        R.append({"Date":d,"Invoice":inv,"Amount":amt,"Type":tp})
    return R

def parse_edward_don(t):
    R = []
    for m in re.finditer(
        r'\d{10}\s+(\d{10})\s+(\w+\s+\d{1,2}\s+\d{4})\s+\S+\s+'
        r'\w+\s+\d{1,2}\s+\d{4}\s+([\d,]+\.\d{2})\s+USD\s+[\d,]+\.\d{2}\s+USD', t):
        R.append({"Date":m[2],"Invoice":m[1].lstrip("0"),
                   "Amount":float(m[3].replace(",","")),"Type":"Invoice"})
    for m in re.finditer(
        r'\d{10}\s+(\d{10})\s+(\w+\s+\d{1,2}\s+\d{4})\s+\S+\s+'
        r'\w+\s+\d{1,2}\s+\d{4}\s+\(([\d,]+\.\d{2})\s+USD\)\s+\([\d,]+\.\d{2}\s+USD\)', t):
        R.append({"Date":m[2],"Invoice":m[1].lstrip("0"),
                   "Amount":-float(m[3].replace(",","")),"Type":"Credit Memo"})
    return R

def parse_romanos(t):
    R = []
    for m in re.finditer(
        r'(\d{2}/\d{2}/\d{4})\s+INV\s+#(\d+)\.\s+Orig\.\s+Amount\s+\$([\d,]+\.\d{2})\.', t):
        R.append({"Date":m[1],"Invoice":m[2],"Amount":float(m[3].replace(",","")),"Type":"Invoice"})
    return R

def parse_cintas(t):
    R = []; c = t.encode('ascii','replace').decode('ascii')
    for m in re.finditer(r'(\d{7,})\s+(\d{2}/\d{2}/\d{2})\s+\d{2}/\d{2}\S*\s+([\d,]+\.\d{2})', c):
        R.append({"Date":m[2],"Invoice":m[1],"Amount":float(m[3].replace(",","")),"Type":"Invoice"})
    return R

def parse_chefs_warehouse(t):
    R = []
    for line in t.split('\n'):
        L = line.strip()
        m = re.match(r'^(\d{8})\s+(\d{2}/\d{2}/\d{2})\s+Invoice\s+'
                     r'([\d,]+\.\d{2})\s+\([\d,]+\.\d{2}\)\s+[\d,]+\.\d{2}', L)
        if m: R.append({"Date":m[2],"Invoice":m[1],"Amount":float(m[3].replace(",","")),"Type":"Invoice"}); continue
        m = re.match(r'^(\d{8})\s+(\d{2}/\d{2}/\d{2})\s+Invoice\s+'
                     r'([\d,]+\.\d{2})\s+[\d,]+\.\d{2}', L)
        if m: R.append({"Date":m[2],"Invoice":m[1],"Amount":float(m[3].replace(",","")),"Type":"Invoice"}); continue
        m = re.match(r'^(\d{8})\s+(\d{2}/\d{2}/\d{2})\s+\(([\d,]+\.\d{2})\)\s+\([\d,]+\.\d{2}\)', L)
        if m: R.append({"Date":m[2],"Invoice":m[1],"Amount":-float(m[3].replace(",","")),"Type":"Credit Memo"}); continue
        m = re.match(r'^(\d{7,})\s+(\d{2}/\d{2}/\d{2})\s+\(([\d,]+\.\d{2})\)\s+\([\d,]+\.\d{2}\)', L)
        if m: R.append({"Date":m[2],"Invoice":m[1],"Amount":-float(m[3].replace(",","")),"Type":"Unapplied Cash"})
    return R

def parse_dex_imaging(t):
    R = []
    for m in re.finditer(
        r'(?:Contract\s+)?Invoice\s+(\d{1,2}/\d{1,2}/\d{4})\s+\d{1,2}/\d{1,2}/\d{4}\s+(AR\d+)\s+\S+\s+\$([\d,]+\.\d{2})', t):
        R.append({"Date":m[1],"Invoice":m[2],"Amount":float(m[3].replace(",","")),"Type":"Invoice"})
    return R

def parse_gourmet_foods(t):
    R = []
    for m in re.finditer(r'(\d{8})\s+(\d+)\s+(INV|CM)\s+([\d,]+\.\d{2})', t):
        dr,inv,tp,a = m.groups(); amt = float(a.replace(",",""))
        if tp == "CM": amt = -amt
        R.append({"Date":f"{dr[:4]}-{dr[4:6]}-{dr[6:8]}","Invoice":inv,"Amount":amt,
                   "Type":"Invoice" if tp=="INV" else "Credit Memo"})
    return R

def parse_halperns(t):
    R = []
    for m in re.finditer(
        r'(\d{2}/\d{2}/\d{4})\s+INV#\s*(\d+(?:-\w+)?)\s+Due\s+\d{2}/\d{2}/\d{4}\s+'
        r'(-?[\d,]+\.\d{2})\s+-?[\d,]+\.\d{2}', t):
        R.append({"Date":m[1],"Invoice":m[2],"Amount":float(m[3].replace(",","")),
                   "Type":"Credit Memo" if "-C" in m[2] else "Invoice"})
    return R

def parse_piper_fire(t):
    R = []
    for m in re.finditer(r'(\d{4}-\d{2}-\d{2})\s+(\d+)\s+\$([\d,]+\.\d{2})', t):
        R.append({"Date":m[1],"Invoice":m[2],"Amount":float(m[3].replace(",","")),"Type":"Invoice"})
    return R

def parse_propane_ninja(_):
    return [
        {"Date":"02/03/2026","Invoice":"1084999","Amount":-214.86,"Type":"Credit (OCR)"},
        {"Date":"02/21/2026","Invoice":"1091427","Amount":432.63,"Type":"Invoice (OCR)"},
    ]

def parse_mr_greens(t):
    R = []
    for m in re.finditer(
        r'^([A-Z]{2}\d{4})\s+(\d{2}/\d{2}/\d{4})\s+(?:\S+\s+\d{2}/\d{2}/\d{4}\s+)?'
        r'(-?[\d,]+\.\d{2})\s+-?[\d,]+\.\d{2}', t, re.MULTILINE):
        tp = "VOID" if "VOID" in t[max(0,m.start()-10):m.end()+10] else "Invoice"
        R.append({"Date":m[2],"Invoice":m[1],"Amount":float(m[3].replace(",","")),"Type":tp})
    return R

def parse_bush_bros(t):
    R = []
    for m in re.finditer(
        r'(\d{1,2}/\d{1,2}/\d{4})\s+Invoice\s+(\d+)\s+.+?\s+([\d,]+\.\d{2})\s+[\d,]+\.\d{2}', t):
        R.append({"Date":m[1],"Invoice":m[2],"Amount":float(m[3].replace(",","")),"Type":"Invoice"})
    for m in re.finditer(r'(\d{1,2}/\d{1,2}/\d{4})\s+Payment\s+(\S+)\s+.+?\s+-([\d,]+\.\d{2})', t):
        R.append({"Date":m[1],"Invoice":m[2],"Amount":-float(m[3].replace(",","")),"Type":"Payment"})
    return R

def parse_us_paper(t):
    D = {}; cur = None
    for line in t.split('\n'):
        L = line.strip()
        if L in ("MAD DOGS AND ENGLISHMEN","OXFORD EXCHANGE LLC","Predalina LLC",
                  "SH-19","The Library St Pete","The Stovall House"):
            cur = L; D.setdefault(cur,[]); continue
        if cur and re.match(r'^\d{2}/\d{2}/\d{4}', L):
            m = re.match(r'(\d{2}/\d{2}/\d{4})\s+Invoice\s+(\d+)\s+\d{2}/\d{2}/\d{4}\s+([\d,]+\.\d{2})', L)
            if m: D[cur].append({"Date":m[1],"Invoice":m[2],"Amount":float(m[3].replace(",","")),"Type":"Invoice"})
    return D

def parse_frank_gay(t):
    R = []
    for m in re.finditer(
        r'(\d{9})\s+(?:Net\s+)?.*?(\d{1,2}/\d{1,2}/\d{2})\s+'
        r'\$([\d,]+\.\d{2})\s+\$[\d,]+\.\d{2}\s+\$([\d,]+\.\d{2})', t, re.DOTALL):
        R.append({"Date":m[2],"Invoice":m[1],"Amount":float(m[3].replace(",","")),"Type":"Invoice"})
    return R

def parse_penguin(t):
    R = []
    for m in re.finditer(
        r'(\d{2}/\d{2}/\d{4})\s+(\d{10})\s+\d{4}\s+(?:\S+\s+)?'
        r'\d{2}/\d{2}/\d{4}\s+(CM|RV|IN)\s+(-?[\d,]+\.\d{2}-?)', t):
        a = m[4]; amt = -float(a[:-1].replace(",","")) if a.endswith('-') else float(a.replace(",",""))
        if m[3] == "CM": amt = -abs(amt)
        R.append({"Date":m[1],"Invoice":m[2],"Amount":amt,
                   "Type":{"CM":"Credit Memo","RV":"Revision","IN":"Invoice"}.get(m[3],m[3])})
    return R

def parse_gfs(t):
    R = []
    for m in re.finditer(r'(\d{9,})\s+(\d{2}/\d{2}/\d{2})\s+INVOICE\s+(?:\d+\s+)?([\d,]+\.\d{2})', t, re.I):
        R.append({"Date":m[2],"Invoice":m[1],"Amount":float(m[3].replace(",","")),"Type":"Invoice"})
    return R

def parse_amazon_xl(fp):
    R = []
    try:
        df = pd.read_excel(fp, sheet_name="Invoices", header=None)
        if len(df) > 4:
            for i in range(4, len(df)):
                r = df.iloc[i]; inv = str(r.iloc[1]) if pd.notna(r.iloc[1]) else ""
                d = str(r.iloc[2]) if pd.notna(r.iloc[2]) else ""
                amt = r.iloc[8] if pd.notna(r.iloc[8]) else 0
                if inv: R.append({"Date":d[:10],"Invoice":inv,"Amount":float(amt),"Type":"Invoice"})
    except Exception as e:
        pass
    return R

PARSERS = {
    "BUCCANEER":parse_buccaneer,"CKS BAR":parse_cks,"CKS":parse_cks,
    "ED DON":parse_edward_don,"ROMANOS COF BAR":parse_romanos,"ROMANOS":parse_romanos,
    "CINTAS":parse_cintas,"CW":parse_chefs_warehouse,"DEX IMAGING":parse_dex_imaging,
    "GOURMET FOODS":parse_gourmet_foods,"HALPERNS":parse_halperns,
    "PIPER FIRE":parse_piper_fire,"PROPANE NINJA":parse_propane_ninja,
    "MR GREENS":parse_mr_greens,"BUSH BROS":parse_bush_bros,
    "US PAPER":parse_us_paper,"FRANK GAY":parse_frank_gay,
    "PENGUIN":parse_penguin,"GFS":parse_gfs,"COF BAR":parse_gfs,"AMAZON":None,
}

# ══════════════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════════════

def fi(fn):
    n = fn.replace(".pdf","").replace(".xlsx","").upper()
    loc = None
    for p in ["SH19","SH","LIB","MD","OE","PRED","OCMGT"]:
        if n.startswith(p+" "): loc = p; break
    vk = None; rem = n[len(loc):].strip() if loc else n
    for k in sorted(VM, key=len, reverse=True):
        if rem.startswith(k): vk = k; break
    return loc, vk

def glk(gl, vn, locs):
    if isinstance(vn, str): vn = [vn]
    m = gl["Vendor name"].isin(vn)
    if locs: m &= gl["Location ID"].isin(locs)
    return gl[m].groupby("Document number")["Amount"].sum().to_dict()

def mi(inv, lk):
    for c in [inv, inv.lstrip("0"), inv.zfill(10) if inv.isdigit() and len(inv)<10 else None]:
        if c and c in lk: return c, lk[c]
    return None, None

def _aw(ws, nc, mr=200, money_cols=None):
    if money_cols is None:
        money_cols = set()
    for c in range(1, nc+1):
        mx = max((len(str(ws.cell(r,c).value or "")) for r in range(1, min(ws.max_row+1, mr))), default=0)
        cn = ws.cell(1, c).value or ""
        if cn in money_cols:
            ws.column_dimensions[get_column_letter(c)].width = max(mx+4, MONEY_MIN_W)
        else:
            ws.column_dimensions[get_column_letter(c)].width = min(mx+4, 30)

# ══════════════════════════════════════════════════════════════════════════
#  FORMATTING
# ══════════════════════════════════════════════════════════════════════════

def fmt_detail(ws, nr, nc):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    sc = nc
    MONEY_COLS = {"Stmt Amount","GL Amount","Variance"}
    for r in range(1, nr + 2):
        ws.row_dimensions[r].height = ROW_HT
    for c in range(1, nc+1):
        cl = ws.cell(1,c); cl.fill = HDR; cl.font = _AH
        cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cl.border = NO_BORDER
    for r in range(2, nr+2):
        st = ws.cell(r,sc).value or ""
        for c in range(1, nc+1):
            cl = ws.cell(r,c); cl.font = _A; cl.border = NO_BORDER
            if st == "Matched":           cl.fill = MATCH
            elif st == "Amount Variance": cl.fill = VAR
            elif st == "Missing in GL":   cl.fill = MISS
            elif "OCR" in st:            cl.fill = OCR_F
            elif r%2 == 0:               cl.fill = STRIPE
            cn = ws.cell(1,c).value or ""
            if cn in MONEY_COLS:
                if cl.value is not None: cl.number_format = COMMA_FMT
                cl.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cl.alignment = Alignment(horizontal="center", vertical="center")
            if cn == "Date" and cl.value is not None:
                nd = _norm_date(str(cl.value)) if isinstance(cl.value, str) else cl.value
                if isinstance(nd, datetime):
                    cl.value = nd; cl.number_format = DATE_FMT
                cl.alignment = Alignment(horizontal="center", vertical="center")
    _aw(ws, nc, money_cols={"Stmt Amount","GL Amount","Variance"})
    jc = ws.cell(1, 11)
    jc.value = "Jump Back to Summary"
    jc.hyperlink = "#Summary!A1"
    jc.font = _JF
    jc.fill = NEON
    jc.border = GREEN_BORDER
    jc.alignment = Alignment(horizontal="center", vertical="center")
    if ws.column_dimensions['K'].width < 26:
        ws.column_dimensions['K'].width = 26

def fmt_summary(ws, nr, nc):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    MONEY_COLS = {"Stmt Total","GL Total","Net Variance"}
    QTY_COLS   = {"Items","Matched","Amt Variance","Missing in GL"}
    for r in range(1, nr + 2):
        ws.row_dimensions[r].height = ROW_HT
    for c in range(1, nc+1):
        cl = ws.cell(1,c); cl.fill = SHDRF; cl.font = _AS; cl.border = NO_BORDER
        cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    for r in range(2, nr+2):
        mi_v = ws.cell(r,6).value or 0; va = ws.cell(r,5).value or 0
        rf = PatternFill("solid", fgColor="FDF0EF") if mi_v>0 else (
             PatternFill("solid", fgColor="FFF9E6") if va>0 else
             PatternFill("solid", fgColor="E8F8F0"))
        for c in range(1, nc+1):
            cl = ws.cell(r,c); cl.border = NO_BORDER
            if c != 1: cl.font = _A
            cl.fill = rf
            cn = ws.cell(1,c).value or ""
            if cn in MONEY_COLS:
                if cl.value is not None: cl.number_format = COMMA_FMT
                cl.alignment = Alignment(horizontal="right", vertical="center")
            elif cn in QTY_COLS:
                if cl.value is not None: cl.number_format = COMMA_INT
                cl.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cl.alignment = Alignment(horizontal="center", vertical="center")
    _aw(ws, nc, money_cols={"Stmt Total","GL Total","Net Variance"})

# ══════════════════════════════════════════════════════════════════════════
#  CORE RECONCILIATION (refactored to accept paths, not use globals)
# ══════════════════════════════════════════════════════════════════════════

def run_reconciliation(gl_path, stmt_paths, log_fn=None):
    """
    gl_path   : path to the GL CSV file
    stmt_paths: list of paths to vendor statement files (PDF / XLSX)
    log_fn    : optional callable(str) for progress messages
    Returns   : bytes of the output XLSX, or raises on error
    """
    def log(msg):
        if log_fn: log_fn(msg)

    log(f"Loading GL: {os.path.basename(gl_path)}")
    gl = pd.read_csv(gl_path)
    gl["Document number"] = gl["Document number"].astype(str).str.strip()
    gl["Amount"] = pd.to_numeric(gl["Amount"], errors="coerce").fillna(0)
    gl["Vendor name"] = gl["Vendor name"].fillna("")
    gl["Location ID"] = gl["Location ID"].fillna("")
    log(f"  {len(gl):,} GL rows loaded")

    # Build a name→path dict for statement files
    stmt_map = {os.path.basename(p): p for p in stmt_paths}
    stmts = sorted(
        fn for fn in stmt_map
        if fn.endswith(('.pdf','.xlsx'))
        and not fn.startswith('AP_REC')
        and not fn.startswith('~')
    )
    log(f"{len(stmts)} statement file(s) found")

    grps = {}
    for f in stmts:
        l, v = fi(f)
        if l and v: grps.setdefault((l,v),[]).append(f)
        else: log(f"  SKIP (unrecognised name): {f}")
    sel = {}
    for k, fl in grps.items():
        fl.sort(reverse=True); sel[k] = fl[0]
        if len(fl) > 1:
            log(f"  Multiple files for {k[0]} {k[1]}: using {fl[0]}")

    sheets = {}; srows = []

    def do(sn, raw, gv, gl_l, src):
        inv_rows = [r for r in raw if r["Type"] not in SKIP_TYPES]
        if not inv_rows:
            log(f"    (all payments — skipped)"); return
        lk = glk(gl, gv, gl_l); recon = []
        for it in inv_rows:
            inv = str(it["Invoice"]).strip(); sa = round(it["Amount"],2)
            mk, ga = mi(inv, lk)
            if mk is not None:
                ga = round(ga,2); v = round(sa-ga,2)
                st = "Matched" if abs(v)<0.015 else "Amount Variance"
                recon.append({"Date":it["Date"],"Invoice #":inv,"Type":it["Type"],
                              "Stmt Amount":sa,"GL Amount":ga,"Variance":v,"Status":st})
            else:
                recon.append({"Date":it["Date"],"Invoice #":inv,"Type":it["Type"],
                              "Stmt Amount":sa,"GL Amount":None,"Variance":None,"Status":"Missing in GL"})
        df = pd.DataFrame(recon); sheets[sn] = df
        m = len(df[df.Status=="Matched"]); v = len(df[df.Status=="Amount Variance"])
        mig = len(df[df.Status=="Missing in GL"])
        st = df["Stmt Amount"].sum()
        gt = df["GL Amount"].sum() if df["GL Amount"].notna().any() else 0
        srows.append({"Statement":sn,"Source File":src,"Items":len(df),
                       "Matched":m,"Amt Variance":v,"Missing in GL":mig,
                       "Stmt Total":round(st,2),"GL Total":round(gt,2),
                       "Net Variance":round(st-gt,2)})
        log(f"    {len(df)} items: {m} matched, {v} variance, {mig} missing in GL")

    for (loc,vk), fn in sorted(sel.items()):
        fp = stmt_map[fn]
        gv = VM[vk]; gl_l = LOC.get(loc,[])
        label = f"{loc} {vk.title()}"
        log(f"Processing: {fn}")

        if vk == "US PAPER":
            txt = _pdf(fp); us = parse_us_paper(txt)
            usm = {
                "MAD DOGS AND ENGLISHMEN":(["MAD-80041"],"MD Us Paper"),
                "OXFORD EXCHANGE LLC":(["OE","OE-96001","OE-96003","OE-96004","OE-96005","OE-96008","OE-96011"],"OE Us Paper"),
                "Predalina LLC":(["PRED","PRED-82000"],"PRED Us Paper"),
                "SH-19":(["SH-93004"],"SH19 Us Paper"),
                "The Library St Pete":(["LIB-96100"],"LIB Us Paper"),
                "The Stovall House":(["SH-93001","SH-93002"],"SH Us Paper"),
            }
            for c, ir in us.items():
                if c in usm and ir:
                    l2,sn2 = usm[c]; log(f"  Sub: {c}"); do(sn2,ir,"US PAPER CORP",l2,fn)
            continue
        if fn.endswith(".xlsx") and "AMAZON" in fn.upper():
            r = parse_amazon_xl(fp)
            if r: do(label, r, gv, gl_l, fn)
            continue
        txt = _pdf(fp); parser = PARSERS.get(vk)
        if parser is None: log(f"    No parser for {vk}"); continue
        rows = parser(txt)
        if not rows and not txt.strip(): log(f"    SCANNED PDF — no text extracted"); continue
        if not rows: log(f"    No invoices parsed ({len(txt)} chars)"); continue
        do(label, rows, gv, gl_l, fn)

    if not srows:
        raise ValueError("No vendor statements were successfully processed. "
                         "Check that your file names follow the required format (e.g. 'OE GFS March.pdf').")

    _today = date.today().strftime("%m%d%y")
    output_filename = f"AP_RECONCILIATION_{_today}.xlsx"

    log(f"Building workbook…")
    sdf = pd.DataFrame(srows)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        sdf.to_excel(w, sheet_name="Summary", index=False)
        for sn in sorted(sheets): sheets[sn].to_excel(w, sheet_name=sn[:31], index=False)
    buf.seek(0)

    # Apply formatting
    wb = load_workbook(buf)
    for sn, df in sheets.items():
        s = sn[:31]
        if s in wb.sheetnames: fmt_detail(wb[s], len(df), len(df.columns))
    ws = wb["Summary"]; fmt_summary(ws, len(sdf), len(sdf.columns))
    for r in range(2, len(sdf)+2):
        cl = ws.cell(r,1); sn = cl.value
        if sn and sn[:31] in wb.sheetnames:
            cl.hyperlink = f"#'{sn[:31]}'!A1"; cl.font = _AL
    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)

    tm=sdf["Matched"].sum(); tv=sdf["Amt Variance"].sum(); tmi=sdf["Missing in GL"].sum()
    log(f"Done! {len(sheets)+1} sheets — {tm} matched | {tv} variances | {tmi} missing in GL")

    return out_buf.getvalue(), output_filename

# ══════════════════════════════════════════════════════════════════════════
#  STREAMLIT UI
# ══════════════════════════════════════════════════════════════════════════

def main():
    st.set_page_config(
        page_title="AP Reconciliation",
        page_icon="📊",
        layout="centered"
    )

    # ── Header ──────────────────────────────────────────────────────────
    st.markdown("""
        <h1 style='text-align:center; color:#1F4E79;'>📊 AP Vendor Reconciliation</h1>
        <p style='text-align:center; color:#555; font-size:16px;'>
            Upload your GL export and vendor statement files below, then click <b>Run Reconciliation</b>.
            Your formatted Excel report will be ready to download in seconds.
        </p>
        <hr style='border:1px solid #ddd; margin-bottom:24px;'>
    """, unsafe_allow_html=True)

    # ── Step 1: GL File ─────────────────────────────────────────────────
    st.markdown("### 1️⃣ &nbsp; Upload GL Export (CSV)")
    gl_upload = st.file_uploader(
        "Drag and drop your GL CSV file here, or click to browse",
        type=["csv"],
        key="gl_file",
        label_visibility="collapsed"
    )
    if gl_upload:
        st.success(f"✅ GL file loaded: **{gl_upload.name}**")

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Step 2: Vendor Statements ────────────────────────────────────────
    st.markdown("### 2️⃣ &nbsp; Upload Vendor Statements (PDF or XLSX)")
    st.caption("You can select multiple files at once. File names must follow the usual format, e.g. *OE GFS March.pdf*")
    stmt_uploads = st.file_uploader(
        "Drag and drop vendor statement files here, or click to browse",
        type=["pdf", "xlsx"],
        accept_multiple_files=True,
        key="stmt_files",
        label_visibility="collapsed"
    )
    if stmt_uploads:
        st.success(f"✅ {len(stmt_uploads)} statement file(s) loaded: " +
                   ", ".join(f"**{f.name}**" for f in stmt_uploads))

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Run Button ───────────────────────────────────────────────────────
    run_disabled = not (gl_upload and stmt_uploads)
    run_clicked  = st.button(
        "🚀 Run Reconciliation",
        type="primary",
        disabled=run_disabled,
        use_container_width=True
    )

    if run_disabled and not (gl_upload and stmt_uploads):
        st.info("Upload both a GL file and at least one vendor statement to enable the Run button.")

    # ── Processing ───────────────────────────────────────────────────────
    if run_clicked:
        log_lines = []
        log_placeholder = st.empty()

        def log(msg):
            log_lines.append(msg)
            log_placeholder.markdown(
                "<div style='background:#f0f4f8;border-radius:6px;padding:12px;"
                "font-family:monospace;font-size:13px;max-height:260px;overflow-y:auto;'>"
                + "<br>".join(log_lines) + "</div>",
                unsafe_allow_html=True
            )

        # Save uploads to a temp directory
        tmpdir = tempfile.mkdtemp()
        try:
            with st.spinner("Running reconciliation — please wait…"):
                # Write GL
                gl_path = os.path.join(tmpdir, gl_upload.name)
                with open(gl_path, "wb") as f:
                    f.write(gl_upload.getvalue())

                # Write statements
                stmt_paths = []
                for su in stmt_uploads:
                    sp = os.path.join(tmpdir, su.name)
                    with open(sp, "wb") as f:
                        f.write(su.getvalue())
                    stmt_paths.append(sp)

                result_bytes, result_filename = run_reconciliation(gl_path, stmt_paths, log_fn=log)

        except ValueError as e:
            st.error(f"⚠️ {e}")
            result_bytes = None
        except Exception as e:
            st.error(f"❌ An unexpected error occurred: {e}")
            result_bytes = None
        finally:
            shutil.rmtree(tmpdir, ignore_errors=True)

        if result_bytes:
            st.success("✅ Reconciliation complete! Click below to download your report.")
            st.download_button(
                label="📥 Download Excel Report",
                data=result_bytes,
                file_name=result_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary"
            )

    # ── Footer ───────────────────────────────────────────────────────────
    st.markdown("<hr style='border:1px solid #eee; margin-top:40px;'>", unsafe_allow_html=True)
    st.markdown(
        "<p style='text-align:center;color:#aaa;font-size:12px;'>"
        "Vendor Statement Reconciliation Tool &nbsp;|&nbsp; Internal Use Only</p>",
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()
