"""
Vendor Statement Reconciliation Engine v6
- Layer 1: Vendor-stated total is authoritative. Always carried through to output.
- Layer 2: GL comparison by invoice number + vendor name.
- Multi-entity statements split into per-entity sheets, each with its own sub-total.
- Layer 1 failures STAY in output, flagged "Manual Review Required" — never excluded.
- Banner at top of each detail sheet shows vendor totals + Layer 1/2 status.
"""
import os, re, io, shutil, tempfile
from datetime import date, datetime
import pandas as pd
import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ENGINE_VERSION = "v6.0-layer1-banner"
print(f"[reconciliation_engine] loaded {ENGINE_VERSION}")

# ── Mappings ─────────────────────────────────────────────────────────────────
LOC = {
    "SH19":["SH-93004"],"SH":["SH-93001","SH-93002"],"LIB":["LIB-96100"],
    "MD":["MAD-80041"],
    "OE":["OE","OE-96001","OE-96003","OE-96004","OE-96005","OE-96008","OE-96011"],
    "PRED":["PRED","PRED-82000"],
    "OCMGT":["OCMGT","OCMGT-71000","OCMGT-71001"],
    "JTS":["JTS-01636","JTS-98001","JTS-98009","JTS-98011","JTS-98012","JTS-98014"],
}
VM = {
    "BUCCANEER":"Buccaneer Linen Service",
    "CKS BAR":"CKS PRODUCE INC","CKS":"CKS PRODUCE INC",
    "ED DON":"Edward Don & Company",
    "ROMANOS COF BAR":"Romanos Bakery Of Boca Inc","ROMANOS":"Romanos Bakery Of Boca Inc",
    "CINTAS":["Cintas","CINTAS CORPORATION (UNIFORMS)","CINTAS CORP"],
    "CW":"The Chefs Warehouse of Florida, LLC","DEX IMAGING":"DEX Imaging LLC",
    "GOURMET FOODS":"GOURMET FOODS INTERNATIONAL LAKELAND INC",
    "HALPERNS":"Halperns's Steak & Seafood","PIPER FIRE":"PIPER FIRE PROTECTION INC",
    "PROPANE NINJA":"Propane Ninja","MR GREENS":"MR GREENS PRODUCE",
    "BUSH BROS":"BUSH BROTHERS PROVISION COMPANY LLC","US PAPER":"US PAPER CORP",
    "AMAZON":"Amazon Capital Services","FRANK GAY":"FRANK GAY SERVICES LLC",
    "PENGUIN":"Penguin Random House LLC","GFS":"Gordon Food Service",
    "COF BAR":"Gordon Food Service",
    "ZWIESEL FORTESSA":"Zwiesel Fortessa Americas LLC",
    "FORTESSA":"Zwiesel Fortessa Americas LLC",
    "UNIFIRST":"Unifirst Corporation",
    "CULIGAN":"Culligan Water","CULLIGAN":"Culligan Water",
    "SAMUELS":"Samuels and Son Seafood South Coast LLC",
    "WRI":"WRIGHTS GOURMET HOUSE",
    "COZZINI":"Cozzini Bros. Inc",
    "NORTON":"W. W. Norton & Company Inc",
    "WW NORTON":"W. W. Norton & Company Inc",
}

# ══════════════════════════════════════════════════════════════════════════════
#  THEME
# ══════════════════════════════════════════════════════════════════════════════
HDR   = PatternFill("solid", fgColor="FF7030")
SHDRF = PatternFill("solid", fgColor="1E1B17")
MATCH = PatternFill("solid", fgColor="1A2B1F")
VAR   = PatternFill("solid", fgColor="2B2510")
MISS  = PatternFill("solid", fgColor="2B1515")
OCR_F = PatternFill("solid", fgColor="1E1B2E")
STRIPE= PatternFill("solid", fgColor="26211C")
ALT   = PatternFill("solid", fgColor="302820")
NEON  = PatternFill("solid", fgColor="FF7030")

_A  = Font(name="Aptos", size=11, color="E8DDD0")
_AH = Font(name="Aptos", size=11, bold=True, color="1E1B17")
_AS = Font(name="Aptos", size=11, bold=True, color="FF7030")
_AL = Font(name="Aptos", size=11, color="FF7030", underline="single")
_JF = Font(name="Aptos", size=11, bold=True,  color="1E1B17")
_AM  = Font(name="Aptos", size=11, color="86EFAC")
_AV  = Font(name="Aptos", size=11, color="FCD34D")
_AMI = Font(name="Aptos", size=11, color="F87171")

THIN_ORANGE = Border(
    left=Side(style="thin", color="FF7030"),
    right=Side(style="thin", color="FF7030"),
    top=Side(style="thin", color="FF7030"),
    bottom=Side(style="thin", color="FF7030"),
)
NO_BORDER = Border()

COMMA_FMT = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
COMMA_INT  = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
DATE_FMT   = 'M/D/YYYY'
ROW_HT     = 18
MONEY_MIN_W= 24

SKIP_TYPES = {"Payment", "Unapplied Cash"}

# ── Date normalizer ──────────────────────────────────────────────────────────
def _norm_date(s):
    if not s or not isinstance(s, str):
        return s
    s = s.strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S",
                "%b %d %Y", "%B %d %Y", "%m/%d/%Y %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return s

# ══════════════════════════════════════════════════════════════════════════════
#  STATEMENT TOTAL FINDER — runs FIRST before any extraction
# ══════════════════════════════════════════════════════════════════════════════

def find_statement_total(text):
    """
    Find the authoritative statement total from vendor statement text.
    Tries patterns in confidence order. Returns (float, label) or (None, None).
    """
    lines = text.split("\n")

    def try_float(s):
        try:
            v = float(str(s).replace(",","").replace("$","").strip())
            return v if v > 0 else None
        except:
            return None

    # P1: Explicit labelled totals — highest confidence
    labelled = [
        (r"Amount\s+Due:\s*([\d,]+\.\d{2})",                                               "Amount Due:"),
        (r"Total[\s\-]+Due\s+([\d,]+\.\d{2})",                                            "Total-Due"),
        (r"Total\s+Due:\s*\$?([\d,]+\.\d{2})",                                            "Total Due:"),
        (r"B[\s]*A[\s]*L[\s]*A[\s]*N[\s]*C[\s]*E[\s]+D[\s]*U[\s]*E[\s]*:.*?([\d,]+\.\d{2})", "BALANCE DUE:"),
        (r"^Total:\s*\$\s*([\d,]+\.\d{2})",                                               "Total: $"),
        (r"^TOTAL\s+\$?([\d,]+\.\d{2})",                                                   "TOTAL"),
        (r"Total:\s+([\d,]+\.\d{2})\s*$",                                                  "Total: EOL"),
        (r"Amount\s+Due\s*\n\s*\$?([\d,]+\.\d{2})",                                     "Amount Due newline"),
        # NEW: end-of-document grand totals (US Paper "TOTAL $X $X", CW "BALANCE DUE 13,450.33")
        (r"^\s*TOTAL\s+\$?([\d,]+\.\d{2})\s+\$?[\d,]+\.\d{2}\s*$",                         "Grand TOTAL line"),
        (r"BALANCE\s+DUE\s+\$?([\d,]+\.\d{2})",                                           "BALANCE DUE"),
        (r"New\s+Balance:?\s*\$?([\d,]+\.\d{2})",                                         "New Balance"),
        (r"Statement\s+Total:?\s*\$?([\d,]+\.\d{2})",                                     "Statement Total"),
        (r"Current\s+Balance:?\s*\$?([\d,]+\.\d{2})",                                     "Current Balance"),
        (r"Outstanding\s+Balance:?\s*\$?([\d,]+\.\d{2})",                                 "Outstanding Balance"),
    ]
    for pat, label in labelled:
        m = re.search(pat, text, re.I | re.M)
        if m:
            v = try_float(m.group(1))
            if v: return v, label

    # P2: CW — "BALANCE 13,450.33" (number AFTER BALANCE keyword, but skip "Starting Balance")
    for m in re.finditer(r"(?<!Starting\s)BALANCE\s+([\d,]+\.\d{2})", text, re.I):
        # Skip if preceded by "Starting" within reasonable distance
        start = max(0, m.start() - 15)
        if "starting" in text[start:m.start()].lower():
            continue
        v = try_float(m.group(1))
        if v: return v, "BALANCE amount"

    # P3: IPR single-invoice — "BALANCE DUE" column header, first $ on data row
    if "BALANCE DUE" in text.upper() and "TOTAL AMOUNT" in text.upper():
        for line in lines:
            if re.search(r"\$[\d,]+\.\d{2}.*\$[\d,]+\.\d{2}.*\$0\.00", line):
                m = re.search(r"\$([\d,]+\.\d{2})", line)
                if m:
                    v = try_float(m.group(1))
                    if v: return v, "IPR BALANCE DUE column"

    # P4: Buccaneer/Romanos — last running balance on last INV# line
    inv_lines = [l for l in lines if re.search(r"INV\s*#\d+", l, re.I)]
    if inv_lines:
        amts = re.findall(r"([\d,]+\.\d{2})\s*$", inv_lines[-1])
        if amts:
            v = try_float(amts[0])
            if v: return v, "last INV running balance"

    # P5: Buddy Brew aging row — last $ value in aging section
    m = re.search(r"\$0\.00\s+\$([\d,]+\.\d{2})\s*$", text, re.M)
    if m:
        v = try_float(m.group(1))
        if v: return v, "aging last value"

    # P6: Aging-row total — line of 5+ amount-like values, last is the total
    # e.g. ".00 .00 .00 749.70 .00 749.70" where the trailing value is TOTAL BALANCE
    # Triggers on any aging-style row regardless of header detection
    for line in lines:
        ls = line.strip()
        tokens = re.findall(r'(?:^|\s)(\.\d{2}|\d[\d,]*\.\d{2})(?=\s|$)', ' ' + ls)
        if len(tokens) >= 5:
            # Aging rows typically have several zero-buckets — sanity check
            zero_like = sum(1 for t in tokens if t in ('.00', '0.00', '0'))
            if zero_like >= 2:
                v = try_float(tokens[-1])
                if v: return v, "aging row last value"

    return None, None


def find_section_totals(text):
    """
    For multi-entity statements (e.g. US Paper), find each section header
    and its sub-total. Returns list of dicts:
      [{"section": "MAD DOGS AND ENGLISHMEN", "total": 2748.37, "start": 234, "end": 567}, ...]
    Sections defined by "Total for <NAME> $X.XX" pattern.
    Returns [] if no sections found.
    """
    results = []
    # Pattern: "Total for <NAME> $X.XX" — name may span multiple lines, total is at the end
    # Strategy: find all "Total for ... $X.XX" markers, then walk backwards to find section start
    pattern = re.compile(
        r"Total\s+for\s+(.+?)\s*\$?([\d,]+\.\d{2})\s*\$?[\d,]*\.?\d*",
        re.I | re.DOTALL
    )

    matches = list(pattern.finditer(text))
    if not matches:
        return []

    # For each match, determine the section name (cleaned) and total
    last_end = 0
    for m in matches:
        raw_name = re.sub(r'\s+', ' ', m.group(1)).strip()
        try:
            total = float(m.group(2).replace(",", ""))
        except:
            continue
        if total <= 0:
            continue
        results.append({
            "section": raw_name,
            "total": total,
            "start": last_end,
            "end": m.end(),
        })
        last_end = m.end()

    return results

# ══════════════════════════════════════════════════════════════════════════════
#  PDF TEXT EXTRACTION
# ══════════════════════════════════════════════════════════════════════════════

def _pdf(fp):
    try:
        pdf = pdfplumber.open(fp)
        t = "\n".join(p.extract_text() or "" for p in pdf.pages)
        pdf.close()
        if t.strip():
            return t
    except Exception:
        pass
    try:
        from pdf2image import convert_from_path
        import pytesseract
        pages = convert_from_path(fp, dpi=200)
        return "\n".join(pytesseract.image_to_string(p) for p in pages)
    except Exception:
        return ""

# ══════════════════════════════════════════════════════════════════════════════
#  UNIVERSAL INVOICE EXTRACTOR
# ══════════════════════════════════════════════════════════════════════════════

def parse_generic(t):
    """
    Universal invoice extractor. Works on any vendor format.
    """
    rows = []
    seen = set()
    inv_lines = set()

    def clean_amt(s):
        s = str(s).replace(',','').replace('$','').strip()
        neg = (s.startswith('(') and s.endswith(')')) or s.startswith('-')
        s = s.strip('()').lstrip('-').strip()
        try:
            v = float(s)
            return -v if neg else v
        except:
            return None

    def is_year(s):
        try: return 1900 <= int(s) <= 2099
        except: return False

    def looks_like_amount(s):
        return bool(re.match(r'^\d{1,6}\.\d{2}$', s))

    def add(inv, amt, date='', typ=None, line_idx=None):
        inv = str(inv).strip().rstrip('.')
        norm = inv.lstrip('0') if re.match(r'^\d+$', inv) else inv
        norm = norm or inv
        if len(norm) < 3: return
        if is_year(norm): return
        if looks_like_amount(norm): return
        if inv in seen or norm in seen: return
        v = clean_amt(amt) if not isinstance(amt, float) else amt
        if v is None or abs(v) < 0.01 or abs(v) > 5_000_000: return
        if typ is None: typ = 'Credit Memo' if v < 0 else 'Invoice'
        seen.add(inv); seen.add(norm)
        if line_idx is not None: inv_lines.add(line_idx)
        rows.append({"Date": date, "Invoice": norm, "Amount": v, "Type": typ})

    lines = t.split('\n')
    DATE = r'(?:\d{1,2}[/\.\-]\d{1,2}[/\.\-]\d{2,4})'

    # S1a: CW format — NUMBER DATE Invoice AMOUNT
    CW_INV = re.compile(rf'^(\d{{7,10}})\s+({DATE})\s+Invoice\s+([\d,]+\.\d{{2}})', re.M | re.I)
    for m in CW_INV.finditer(t):
        line_no = t[:m.start()].count('\n')
        add(m.group(1), m.group(3), date=m.group(2), typ='Invoice', line_idx=line_no)

    # S1b: CW credit — NUMBER DATE (AMOUNT) with prev line = Credit Memo
    CW_CM = re.compile(rf'^(\d{{7,10}})\s+({DATE})\s+(\(\d[\d,]*\.\d{{2}}\))', re.M)
    for m in CW_CM.finditer(t):
        line_no = t[:m.start()].count('\n')
        if line_no in inv_lines: continue
        prev = [l for l in lines[:line_no] if l.strip()]
        if prev and 'credit memo' in prev[-1].lower():
            add(m.group(1), m.group(3), date=m.group(2), typ='Credit Memo', line_idx=line_no)

    # S2: Explicit INV/Invoice label — "Invoice #97128:", "INV# 12069891", "INV-70432"
    INV_PAT = re.compile(
        r'(?:Invoice|INV)\s*[#.\-]+\s*(?:INV[#\-\s]?)?([A-Z]?\d[A-Z0-9\-]*)'
        r'.*?\$?\s*(-?\s*\(?\d[\d,]*\.\d{2}\)?)', re.I)
    CM_PAT = re.compile(
        r'Credit[_ ]?Memo[#\s\.\-]+([A-Z]?\d[A-Z0-9\-]*)'
        r'.*?\$?\s*(-?\(?\d[\d,]*\.\d{2}\)?)', re.I)

    for i, line in enumerate(lines):
        if i in inv_lines: continue
        for m in INV_PAT.finditer(line):
            inv_id = m.group(1).rstrip('.')
            if not looks_like_amount(inv_id):
                add(inv_id, m.group(2), line_idx=i)
                inv_lines.add(i)
                break
        if i not in inv_lines:
            m = CM_PAT.search(line)
            if m and not looks_like_amount(m.group(1)):
                v = clean_amt(m.group(2))
                if v: add(m.group(1), -abs(v), typ='Credit Memo', line_idx=i)

    # S3: US Paper — DATE Invoice/Credit_Memo NUMBER DATE AMOUNT
    USP_PAT = re.compile(
        rf'({DATE})\s+(Invoice|Credit\s*Memo)\s+(\d+)\s+{DATE}\s+'
        r'(-?[\d,]+\.\d{2})', re.I)
    for m in USP_PAT.finditer(t):
        line_no = t[:m.start()].count('\n')
        if line_no in inv_lines: continue
        typ = 'Credit Memo' if 'credit' in m.group(2).lower() else 'Invoice'
        v = clean_amt(m.group(4))
        if typ == 'Credit Memo' and v and v > 0: v = -v
        if v is not None: add(m.group(3), str(v), date=m.group(1), typ=typ, line_idx=line_no)

    # S4: GFS-style — long number + date + Invoice/Credit + amount
    # Handles all GFS variants:
    #   "9033865171 03/30/26 INVOICE 787.66"                 (plain, no $)
    #   "924113414 04/08/26 INVOICE 924113414 48.46"         (with PO column)
    #   "2003339809 04/20/26 CREDIT MEMO 9034593368 8.15-"   (credit memo, trailing minus)
    #   "9033821167 03/28/26 Invoice $1,111.47"              (legacy with $)
    GFS_PAT = re.compile(
        rf'^\s*(\d{{6,12}})\s+({DATE})\s+(Invoice|Credit\s*Memo|Credit)\b', re.I | re.M)
    for m in GFS_PAT.finditer(t):
        inv = m.group(1)
        line_no = t[:m.start()].count('\n')
        if line_no in inv_lines or is_year(inv): continue
        line_end = t.find('\n', m.end())
        rest = t[m.end():line_end] if line_end != -1 else t[m.end():]

        # Try $-prefixed amounts first (legacy GFS format)
        dollar_amts = re.findall(r'(-?\s*\$\s*)(\(?)(\d[\d,]+\.\d{2})(\)?)', rest)
        v = None
        if dollar_amts:
            prefix, op, digits, cp = dollar_amts[-1]
            v = float(digits.replace(',', ''))
            if '-' in prefix or (op and cp): v = -v
        else:
            # Plain numeric amount, possibly with trailing minus (credit memo) or PO column before
            # Take the LAST plain amount on the line.
            m2 = re.search(r'(-?[\d,]+\.\d{2})\s*(-?)\s*$', rest)
            if m2:
                v = float(m2.group(1).replace(',', '').lstrip('-'))
                if m2.group(1).startswith('-') or m2.group(2) == '-':
                    v = -v
        if v is None: continue

        typ = 'Credit Memo' if 'credit' in m.group(3).lower() else 'Invoice'
        if typ == 'Credit Memo' and v > 0: v = -v
        add(inv, v, date=m.group(2), typ=typ, line_idx=line_no)

    # S8: Norton/ledger — DATE NUMBER PO AMOUNT CRD/INV type at end
    # "8/12/25 12/10/25 3007785 5844 140.44 INV"
    # "11/10/25 12/10/25 3271161 VR6204 160.40- CRD"
    NORTON_PAT = re.compile(
        rf'({DATE})\s+{DATE}\s+(\d{{7,}})\s+\S+\s+(\d[\d,]*\.\d{{2}})(-?)\s+(INV|CRD|CR)', re.I)
    for m in NORTON_PAT.finditer(t):
        line_no = t[:m.start()].count('\n')
        if line_no in inv_lines: continue
        v = float(m.group(3).replace(',',''))
        is_credit = m.group(5).upper() in ('CRD','CR') or m.group(4) == '-'
        if is_credit: v = -v
        typ = 'Credit Memo' if is_credit else 'Invoice'
        add(m.group(2), v, date=m.group(1), typ=typ, line_idx=line_no)

    # S5: Hachette/aging — [code] DATE 7-12digit ... amounts
    for i, line in enumerate(lines):
        if i in inv_lines: continue
        m = re.search(rf'(?:^|\s)({DATE})\s+(\d{{7,12}})\b', line.strip())
        if m and not is_year(m.group(2)):
            amts = re.findall(r'(?<!\d)(\d[\d,]*\.\d{2})(?!\d)', line)
            if amts:
                add(m.group(2), amts[-1], date=m.group(1), line_idx=i)

    # S6: Edward Don — 10-digit IDs + month date year + USD
    EDDON_PAT = re.compile(
        r'\d{10}\s+(\d{10})\s+(\w+\s+\d{1,2}\s+\d{4})\s+'
        r'(?:\S+\s+)?\w+\s+\d{1,2}\s+\d{4}\s+'
        r'(\(?\d[\d,]+\.\d{2}\)?)\s+USD', re.I)
    for m in EDDON_PAT.finditer(t):
        add(m.group(1).lstrip('0'), m.group(3), date=m.group(2))

    # S7: Samuels/ledger — DATE TYPE NUMBER [PO#] AMOUNT[-]
    SAM_PAT = re.compile(rf'({DATE})\s+([ICPicp])\s+(\d{{5,}})(?:\s+\d+)?\s+(\d[\d,]*\.\d{{2}})(-?)')
    for m in SAM_PAT.finditer(t):
        letter = m.group(2).upper()
        v = float(m.group(4).replace(',',''))
        if m.group(5) == '-' or letter in ('C','P'): v = -v
        typ = 'Invoice' if letter == 'I' else 'Credit Memo'
        line_no = t[:m.start()].count('\n')
        if line_no not in inv_lines:
            add(m.group(3), v, date=m.group(1), typ=typ, line_idx=line_no)

    # S7b: Samuels CASH payment
    CASH_PAT = re.compile(rf'({DATE})\s+P\s+CASH\s+(\d[\d,]*\.\d{{2}})-')
    cash_n = 0
    for m in CASH_PAT.finditer(t):
        cash_n += 1
        line_no = t[:m.start()].count('\n')
        if line_no not in inv_lines:
            add(f'CASH{cash_n}', f'-{m.group(2)}', date=m.group(1), typ='Credit Memo', line_idx=line_no)

    # S10: CKS-style aging detail — "Invoice DATE INV#### DATE age $amount"
    # Format: "Invoice 03/31/2026 INV4869784 04/21/2026 16 $293.90"
    # Also handles: "Payment 04/06/2026 PMT6816872 04/06/2026 10 ($244.67)"
    AGING_PAT = re.compile(
        rf'^\s*(Invoice|Credit\s*Memo|Payment)\s+({DATE})\s+'
        rf'([A-Z]+\d{{4,}})\s+{DATE}\s+\d+\s+'
        r'\(?\$?(-?[\d,]+\.\d{2})\)?\s*$', re.M | re.I)
    for m in AGING_PAT.finditer(t):
        line_no = t[:m.start()].count('\n')
        if line_no in inv_lines: continue
        type_word = m.group(1)
        v_str = m.group(4)
        # Check if amount was wrapped in parens (= negative)
        full_match = m.group(0)
        is_paren_neg = '(' in full_match and ')' in full_match[full_match.index(v_str):]
        v = float(v_str.replace(',', '').lstrip('-'))
        if is_paren_neg or v_str.startswith('-'):
            v = -v
        # Determine type
        tw = type_word.lower()
        if 'payment' in tw:
            typ = 'Payment'
            if v > 0: v = -v  # payments reduce balance
        elif 'credit' in tw:
            typ = 'Credit Memo'
            if v > 0: v = -v
        else:
            typ = 'Invoice'
        add(m.group(3), v, date=m.group(2), typ=typ, line_idx=line_no)

    # S11: IN-prefix doc — "IN3760500 4/6/2026 IN 2026-04-06 ... 5/6/2026 1,320.56"
    # Doc number is letters+digits, line ends with the amount
    INPREFIX_PAT = re.compile(
        rf'^\s*([A-Z]{{2,4}}\d{{4,}})\s+({DATE})\s+([A-Z]{{2,3}})\b.*?'
        r'\b(-?\(?[\d,]+\.\d{2}\)?)\s*$', re.M)
    for m in INPREFIX_PAT.finditer(t):
        line_no = t[:m.start()].count('\n')
        if line_no in inv_lines: continue
        typ_code = m.group(3).upper()
        v_str = m.group(4).strip()
        is_neg = (v_str.startswith('(') and v_str.endswith(')')) or v_str.startswith('-')
        v = float(v_str.strip('()').lstrip('-').replace(',', ''))
        if is_neg: v = -v
        if typ_code in ('CR', 'CM', 'CN'):
            typ = 'Credit Memo'
            if v > 0: v = -v
        elif typ_code in ('PY', 'UC', 'PMT'):
            typ = 'Payment'
        else:
            typ = 'Invoice'
        add(m.group(1), v, date=m.group(2), typ=typ, line_idx=line_no)

    # S12: Arcadia-style — "DATE TYPE DOC# REF DATE Amount1 Amount2 Amount3"
    # Capture the MIDDLE amount (Outstanding), not first (Original) or last (Balance)
    ARCADIA_PAT = re.compile(
        rf'^\s*({DATE})\s+(INV|CR|CM|CRD|DB)\s+(\d{{5,}})\s+\S+\s+'
        rf'{DATE}\s+([\d,]+\.\d{{2}})\s+([\d,]+\.\d{{2}})\s+([\d,]+\.\d{{2}})\s*$',
        re.M | re.I)
    for m in ARCADIA_PAT.finditer(t):
        line_no = t[:m.start()].count('\n')
        if line_no in inv_lines: continue
        typ_code = m.group(2).upper()
        v = float(m.group(5).replace(',', ''))  # AMOUNT OUTSTANDING (middle)
        if typ_code in ('CR', 'CM', 'CRD'):
            typ = 'Credit Memo'
            if v > 0: v = -v
        else:
            typ = 'Invoice'
        add(m.group(3), v, date=m.group(1), typ=typ, line_idx=line_no)

    # S13: JLA-style alphanumeric reference — "DATE DATE alphanum-ref ... Amount Balance"
    # "3/16/26 3/16/26 m128554 PO# ESPRESSO 28,880.09 28,880.09"
    JLA_PAT = re.compile(
        rf'^\s*({DATE})\s+{DATE}\s+([a-zA-Z]+\d{{3,}})\s+\S+.*?'
        r'\b(\d[\d,]*\.\d{2})\s+\d[\d,]*\.\d{2}\s*$', re.M)
    for m in JLA_PAT.finditer(t):
        line_no = t[:m.start()].count('\n')
        if line_no in inv_lines: continue
        v = float(m.group(3).replace(',', ''))
        add(m.group(2), v, date=m.group(1), typ='Invoice', line_idx=line_no)

    # S14: No invoice number — "DATE Type $Amount $Amount $Amount" (e.g. AVIT)
    # Synthesize invoice key from date so the row appears in output (Layer 1 still validates).
    # Only run if very few rows have been extracted so far (avoid noise on rich statements).
    if len(rows) < 2:
        NOINV_PAT = re.compile(
            rf'^\s*({DATE})\s+(Invoice|Credit\s*Memo|Payment)\s+'
            r'\$([\d,]+\.\d{2})\s+\$[\d,]+\.\d{2}', re.M | re.I)
        for m in NOINV_PAT.finditer(t):
            line_no = t[:m.start()].count('\n')
            if line_no in inv_lines: continue
            tw = m.group(2).lower()
            v = float(m.group(3).replace(',', ''))
            if 'credit' in tw:
                typ = 'Credit Memo'
                if v > 0: v = -v
            elif 'payment' in tw:
                typ = 'Payment'
                if v > 0: v = -v
            else:
                typ = 'Invoice'
            # Synthetic invoice key — clearly artificial so it lands as "Missing in GL"
            synth_inv = f"STMT-{m.group(1).replace('/', '')}"
            add(synth_inv, v, date=m.group(1), typ=typ, line_idx=line_no)

    # S9: Last resort — $amount lines not yet processed (runs LAST so specific strategies win)
    for i, line in enumerate(lines):
        if i in inv_lines: continue
        if re.search(r'\$[\d,]+\.\d{2}', line):
            nums = re.findall(r'\b(\d{4,12})\b', line)
            amts = re.findall(r'\$([\d,]+\.\d{2})', line)
            if nums and amts:
                for n in nums:
                    if not is_year(n) and not looks_like_amount(n) and n not in seen:
                        add(n, amts[0], line_idx=i); break

    return rows


def parse_wrights(t):
    R = []
    inv_lines = re.findall(r'(\d{2}/\d{2}/\d{2})\s+(\d{7})', t)
    parts = re.split(r'Amount\s+Balance\s+Due\s+by', t, maxsplit=1, flags=re.IGNORECASE)
    amt_section = parts[1] if len(parts) > 1 else t
    amt_lines = re.findall(r'([\d,]+\.\d{2})\s+[\d,]+\.\d{2}\s+\d{2}/\d{2}/\d{2}', amt_section)
    for i, (date, inv) in enumerate(inv_lines):
        if i < len(amt_lines):
            R.append({"Date":date,"Invoice":inv,
                      "Amount":float(amt_lines[i].replace(",","")),"Type":"Invoice"})
    return R


def parse_us_paper(t):
    D = {}; cur = None
    for line in t.split('\n'):
        L = line.strip()
        if L in ("MAD DOGS AND ENGLISHMEN","OXFORD EXCHANGE LLC","Predalina LLC",
                  "SH-19","The Library St Pete","The Stovall House"):
            cur = L; D.setdefault(cur,[]); continue
        if cur and re.match(r'^\d{2}/\d{2}/\d{4}', L):
            m = re.match(r'(\d{2}/\d{2}/\d{4})\s+(Invoice|Credit\s*Memo)\s+(\d+)\s+\d{2}/\d{2}/\d{4}\s+(-?[\d,]+\.\d{2})', L)
            if m:
                v = float(m.group(4).replace(',',''))
                if 'credit' in m.group(2).lower() and v > 0: v = -v
                D[cur].append({"Date":m.group(1),"Invoice":m.group(3),"Amount":v,
                               "Type":"Credit Memo" if v<0 else "Invoice"})
    return D


def parse_amazon_xl(fp):
    R = []
    try:
        df = pd.read_excel(fp, sheet_name="Invoices", header=None)
        if len(df) > 4:
            for i in range(4, len(df)):
                r = df.iloc[i]; inv = str(r.iloc[1]) if pd.notna(r.iloc[1]) else ""
                d = str(r.iloc[2]) if pd.notna(r.iloc[2]) else ""
                amt = r.iloc[8] if pd.notna(r.iloc[8]) else 0
                if inv: R.append({"Date":d[:10],"Invoice":inv,"Amount":float(amt),"Type":"Invoice"})
    except Exception:
        pass
    return R


# All vendors route through parse_generic now
PARSERS = {k: parse_generic for k in VM if k not in ("US PAPER", "WRI", "AMAZON")}
PARSERS["US PAPER"] = parse_us_paper
PARSERS["WRI"] = parse_wrights
PARSERS["COF BAR"] = parse_generic

# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def fi(fn):
    # Normalize: replace underscores with spaces (some uploads convert spaces → underscores)
    n = fn.replace(".pdf","").replace(".xlsx","").upper()
    n = n.replace("_", " ")
    n = re.sub(r'\s+', ' ', n).strip()
    loc = None
    for p in ["SH19","SH","LIB","MD","OE","PRED","OCMGT","JTS"]:
        if n.startswith(p+" "): loc = p; break
    vk = None; rem = n[len(loc):].strip() if loc else n
    for k in sorted(VM, key=len, reverse=True):
        if rem.startswith(k): vk = k; break
    return loc, vk

def glk(gl, vn, locs):
    if isinstance(vn, str): vn = [vn]
    m = gl["Vendor name"].isin(vn)
    if locs: m &= gl["Location ID"].isin(locs)
    return gl[m].groupby("Document number")["Amount"].sum().to_dict()

def mi(inv, lk):
    import re as _re
    inv = str(inv).strip()
    candidates = [
        inv,
        inv.lstrip("0") or inv,
        inv.zfill(10) if inv.isdigit() and len(inv)<10 else None,
        f"INV-{inv}" if inv.isdigit() else None,
        f"INV{inv}" if inv.isdigit() else None,
    ]
    # Strip ANY leading alpha prefix (INV, IN, PMT, CR, CM, etc.) and also try lstripped
    m = _re.match(r'^([A-Z]+)[-\s]?(\d+)$', inv, _re.I)
    if m:
        num = m.group(2)
        candidates.extend([num, num.lstrip("0") or num])
    for c in candidates:
        if c and c in lk: return c, lk[c]
    return None, None

def _aw(ws, nc, mr=200, money_cols=None):
    if money_cols is None:
        money_cols = set()
    for c in range(1, nc+1):
        mx = max((len(str(ws.cell(r,c).value or "")) for r in range(1, min(ws.max_row+1, mr))), default=0)
        cn = ws.cell(1, c).value or ""
        if cn in money_cols:
            ws.column_dimensions[get_column_letter(c)].width = max(mx+4, MONEY_MIN_W)
        else:
            ws.column_dimensions[get_column_letter(c)].width = min(mx+4, 30)

# ══════════════════════════════════════════════════════════════════════════════
#  FORMATTING
# ══════════════════════════════════════════════════════════════════════════════

def _write_banner(ws, meta, banner_rows):
    """Write the metadata banner at the top of a detail sheet (rows 1..banner_rows)."""
    LBL = Font(name="Aptos", size=11, bold=True, color="FF7030")
    VAL = Font(name="Aptos", size=11, color="E8DDD0")
    OK  = Font(name="Aptos", size=11, bold=True, color="86EFAC")
    WRN = Font(name="Aptos", size=11, bold=True, color="FCD34D")
    ERR = Font(name="Aptos", size=11, bold=True, color="F87171")

    BG = PatternFill("solid", fgColor="1E1B17")

    def status_font(status):
        if not status: return VAL
        s = str(status)
        if s.startswith("✓"): return OK
        if s.startswith("⚠"): return WRN
        if s.startswith("❓"): return ERR
        return VAL

    def fmt_money(v):
        if v is None: return "—"
        return f"${v:,.2f}"

    def fmt_var(v):
        if v is None: return "—"
        return f"${v:+,.2f}"

    rows = [
        ("Vendor:",              meta.get("vendor_display", "")),
        ("Entity:",              meta.get("entity_display", "")),
        ("Statement Date:",      meta.get("stmt_date", "") or "—"),
        ("Source File:",         meta.get("source_file", "")),
        ("",                     ""),
        ("Vendor Stated Total:", fmt_money(meta.get("vendor_stated_total"))),
        ("Extracted Total:",     fmt_money(meta.get("extracted_total"))),
        ("Layer 1 Variance:",    fmt_var(meta.get("l1_variance"))),
        ("Layer 1 Status:",      meta.get("l1_status", "")),
        ("",                     ""),
        ("GL Total (Matched):",  fmt_money(meta.get("gl_total"))),
        ("Net Variance:",        fmt_var(meta.get("net_variance"))),
        ("Layer 2 Status:",      meta.get("l2_status", "")),
    ]

    for i, (label, value) in enumerate(rows, start=1):
        lc = ws.cell(row=i, column=1, value=label)
        vc = ws.cell(row=i, column=2, value=value)
        lc.fill = BG; vc.fill = BG
        lc.font = LBL
        if label.startswith("Layer 1 Status") or label.startswith("Layer 2 Status"):
            vc.font = status_font(value)
        elif label.startswith("Layer 1 Variance") or label.startswith("Net Variance"):
            # Color variance amounts: green if 0, yellow if non-zero
            try:
                v = meta.get("l1_variance" if "Layer 1" in label else "net_variance")
                if v is None or abs(float(v)) <= 0.02:
                    vc.font = OK
                else:
                    vc.font = WRN
            except:
                vc.font = VAL
        else:
            vc.font = VAL
        lc.alignment = Alignment(horizontal="right", vertical="center")
        vc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[i].height = ROW_HT

    # Fill remaining cells in banner area with BG so it looks intentional
    for i in range(1, banner_rows + 1):
        for c in range(3, 12):
            cell = ws.cell(row=i, column=c)
            cell.fill = BG
            cell.border = NO_BORDER

    # Set widths so banner labels and values are readable
    ws.column_dimensions['A'].width = max(ws.column_dimensions['A'].width or 0, 24)
    ws.column_dimensions['B'].width = max(ws.column_dimensions['B'].width or 0, 40)


def fmt_detail(ws, nr, nc, banner_rows=0):
    """Format the detail data table. Table header is at row banner_rows+1, data follows."""
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "FF7030"
    header_row = banner_rows + 1
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    MONEY_COLS = {"Stmt Amount","GL Amount","Variance"}

    for r in range(header_row, header_row + nr + 1):
        ws.row_dimensions[r].height = ROW_HT

    # Header row
    for c in range(1, nc+1):
        cl = ws.cell(header_row, c)
        cl.fill = HDR; cl.font = _AH; cl.border = NO_BORDER
        cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Data rows
    for r in range(header_row + 1, header_row + nr + 1):
        st = ws.cell(r, nc).value or ""
        for c in range(1, nc+1):
            cl = ws.cell(r, c); cl.border = NO_BORDER
            cn = ws.cell(header_row, c).value or ""

            if st == "Matched":
                cl.fill = MATCH
                cl.font = _AM if cn in MONEY_COLS else _A
            elif st == "Amount Variance":
                cl.fill = VAR
                cl.font = _AV if cn in MONEY_COLS else _A
            elif st == "Missing in GL":
                cl.fill = MISS
                cl.font = _AMI if cn in MONEY_COLS else _A
            elif "OCR" in st:
                cl.fill = OCR_F; cl.font = _A
            elif r % 2 == 0:
                cl.fill = STRIPE; cl.font = _A
            else:
                cl.fill = ALT; cl.font = _A

            if cn in MONEY_COLS:
                if cl.value is not None: cl.number_format = COMMA_FMT
                cl.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cl.alignment = Alignment(horizontal="center", vertical="center")

            if cn == "Date" and cl.value is not None:
                nd = _norm_date(str(cl.value)) if isinstance(cl.value, str) else cl.value
                if isinstance(nd, datetime):
                    cl.value = nd; cl.number_format = DATE_FMT
                cl.alignment = Alignment(horizontal="center", vertical="center")

    _aw_offset(ws, nc, header_row=header_row, mr=banner_rows + nr + 200, money_cols=MONEY_COLS)

    # Back-link button next to the table header
    jc = ws.cell(header_row, 11)
    jc.value = "\u2190 Summary"
    jc.hyperlink = "#Summary!A1"
    jc.font = _JF
    jc.fill = NEON
    jc.border = NO_BORDER
    jc.alignment = Alignment(horizontal="center", vertical="center")
    if (ws.column_dimensions['K'].width or 0) < 20:
        ws.column_dimensions['K'].width = 20


def _aw_offset(ws, nc, header_row=1, mr=200, money_cols=None):
    """Auto-width that uses header row at given offset (for banner-aware layout)."""
    if money_cols is None:
        money_cols = set()
    for c in range(1, nc+1):
        # Check both banner and table content for max width
        mx = max((len(str(ws.cell(r, c).value or ""))
                  for r in range(1, min(ws.max_row + 1, mr))), default=0)
        cn = ws.cell(header_row, c).value or ""
        existing = ws.column_dimensions[get_column_letter(c)].width or 0
        if cn in money_cols:
            new_w = max(mx + 4, MONEY_MIN_W)
        else:
            new_w = min(mx + 4, 40)
        ws.column_dimensions[get_column_letter(c)].width = max(existing, new_w)


def fmt_summary(ws, nr, nc):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "FF7030"
    ws.freeze_panes = "A2"
    MONEY_COLS = {"Vendor Stmt Total","Extracted Total","Layer 1 Variance",
                  "GL Total","Net Variance"}
    QTY_COLS   = {"Items","Matched","Amt Variance","Missing in GL"}
    STATUS_COLS = {"Layer 1 Status","Layer 2 Status"}

    for r in range(1, nr + 2):
        ws.row_dimensions[r].height = ROW_HT

    # Header row
    for c in range(1, nc+1):
        cl = ws.cell(1, c)
        cl.fill = SHDRF; cl.font = _AS; cl.border = NO_BORDER
        cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    OK  = Font(name="Aptos", size=11, bold=True, color="86EFAC")
    WRN = Font(name="Aptos", size=11, bold=True, color="FCD34D")
    ERR = Font(name="Aptos", size=11, bold=True, color="F87171")

    # Locate column indexes by header name
    headers = {ws.cell(1, c).value: c for c in range(1, nc+1)}
    l1_status_col = headers.get("Layer 1 Status")
    l2_status_col = headers.get("Layer 2 Status")
    mig_col       = headers.get("Missing in GL")
    var_col       = headers.get("Amt Variance")

    for r in range(2, nr+2):
        l1_status = ws.cell(r, l1_status_col).value if l1_status_col else ""
        l2_status = ws.cell(r, l2_status_col).value if l2_status_col else ""
        mi_v      = ws.cell(r, mig_col).value or 0 if mig_col else 0
        va        = ws.cell(r, var_col).value or 0 if var_col else 0

        # Row fill = worst status across both layers
        l1_bad = isinstance(l1_status, str) and l1_status.startswith("⚠")
        l1_unk = isinstance(l1_status, str) and l1_status.startswith("❓")
        l2_bad = mi_v > 0 or va > 0
        if l1_bad or l1_unk or mi_v > 0:
            rf = MISS
        elif l2_bad:
            rf = VAR
        else:
            rf = MATCH

        for c in range(1, nc+1):
            cl = ws.cell(r, c); cl.border = NO_BORDER
            cl.fill = rf
            cn = ws.cell(1, c).value or ""

            if cn in MONEY_COLS:
                if mi_v > 0 or l1_bad:
                    cl.font = _AMI
                elif l2_bad or l1_unk:
                    cl.font = _AV
                else:
                    cl.font = _AM
                if cl.value is not None: cl.number_format = COMMA_FMT
                cl.alignment = Alignment(horizontal="right", vertical="center")
            elif cn in QTY_COLS:
                cl.font = _A
                if cl.value is not None: cl.number_format = COMMA_INT
                cl.alignment = Alignment(horizontal="center", vertical="center")
            elif cn in STATUS_COLS:
                v = cl.value or ""
                if isinstance(v, str) and v.startswith("✓"):
                    cl.font = OK
                elif isinstance(v, str) and v.startswith("⚠"):
                    cl.font = WRN
                elif isinstance(v, str) and v.startswith("❓"):
                    cl.font = ERR
                else:
                    cl.font = _A
                cl.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cl.font = _A
                cl.alignment = Alignment(horizontal="center", vertical="center")

    _aw(ws, nc, money_cols=MONEY_COLS)

# ══════════════════════════════════════════════════════════════════════════════
#  CORE RECONCILIATION
# ══════════════════════════════════════════════════════════════════════════════

def smart_invoice_match(raw_rows, gl, log_fn=None):
    def log(m):
        if log_fn: log_fn(m)

    inv_rows = [r for r in raw_rows if r.get("Type","") not in SKIP_TYPES]
    if not inv_rows:
        return []

    import re as _re

    def _gen_variants(raw):
        """Generate all reasonable invoice-number variants for matching against GL."""
        raw = str(raw).strip()
        out = [raw, raw.lstrip("0") or raw]
        if raw.isdigit():
            if len(raw) < 10:
                out.append(raw.zfill(10))
            out.extend([f"INV-{raw}", f"INV{raw}"])
        # Strip any leading alpha prefix (INV, IN, PMT, CR, CM, etc.)
        m = _re.match(r'^([A-Z]+)[-\s]?(\d+)$', raw, _re.I)
        if m:
            num = m.group(2)
            out.extend([num, num.lstrip("0") or num])
        return [v for v in out if v]

    variants = {}
    for r in inv_rows:
        raw = str(r["Invoice"]).strip()
        for v in _gen_variants(raw):
            variants[v] = raw

    gl_hit = gl[gl["Document number"].isin(variants.keys())].copy()

    if gl_hit.empty:
        log("  smart-match: no invoice numbers found in GL")
        return []

    log(f"  smart-match: {len(gl_hit)} GL entries matched from {len(inv_rows)} invoices")

    results = []
    for (vendor, loc_id), grp in gl_hit.groupby(["Vendor name","Location ID"]):
        gl_docs = set(grp["Document number"].tolist())
        sub = [r for r in inv_rows
               if any(v in gl_docs for v in _gen_variants(r["Invoice"]))]

        if sub:
            total_inv = len(inv_rows)
            if total_inv <= 2:
                min_match = 1
            else:
                min_match = max(2, int(total_inv * 0.20))
            if len(sub) < min_match:
                log(f"    ↳ skipping {vendor} / {loc_id}: {len(sub)}/{total_inv} (below threshold)")
                continue
            # Build a readable label — skip stop words like "The", "A", "An", "Of"
            STOPS = {"THE", "A", "AN", "OF", "&"}
            words = [w for w in vendor.split() if w.upper() not in STOPS]
            short_name = (words[0] if words else vendor.split()[0]).rstrip(',.')
            label = f"{loc_id} {short_name}"[:31]
            log(f"    → {vendor} / {loc_id}: {len(sub)} invoices")
            results.append({"label": label, "rows": sub,
                             "vendor": vendor, "loc_id": loc_id})

    return results


def run_reconciliation(gl_path, stmt_paths, log_fn=None, file_overrides=None):
    """
    gl_path   : path to the GL CSV file
    stmt_paths: list of paths to vendor statement files (PDF / XLSX)
    log_fn    : optional callable(str) for progress messages
    Returns   : (bytes, filename, reconciled_set, skipped_list)
    """
    def log(msg):
        if log_fn: log_fn(msg)

    log(f"Loading GL: {os.path.basename(gl_path)}")
    gl = pd.read_csv(gl_path)
    gl["Document number"] = gl["Document number"].astype(str).str.strip()
    gl["Amount"] = pd.to_numeric(gl["Amount"], errors="coerce").fillna(0)
    gl["Vendor name"] = gl["Vendor name"].fillna("")
    gl["Location ID"] = gl["Location ID"].fillna("")
    log(f"  {len(gl):,} GL rows loaded")

    stmt_map = {os.path.basename(p): p for p in stmt_paths}
    stmts = sorted(
        fn for fn in stmt_map
        if fn.endswith(('.pdf','.xlsx'))
        and not fn.startswith('AP_REC')
        and not fn.startswith('~')
    )
    log(f"{len(stmts)} statement file(s) found")

    all_stmt_set = set(stmts)
    sheets = {}; sheet_meta = {}; srows = []; reconciled = set()

    if file_overrides is None:
        file_overrides = {}

    def do(sn, raw, gv, gl_l, src,
           vendor_stated_total=None, vendor_display=None,
           entity_display=None, stmt_date=None):
        """
        Build one sheet with reconciliation results + Layer 1/2 metadata.
        - If raw is empty → creates a stub sheet flagged "Extraction Failed".
        - Payments / Unapplied Cash appear on sheet but are NOT reconciled (per spec).
        - Layer 1 failures STAY in output; never excluded.
        """
        base_sn = sn[:31]; sn2 = base_sn; n = 1
        while sn2 in sheets:
            sn2 = f"{base_sn[:28]} {n:02d}"; n += 1
        sn = sn2

        # ── Extraction failed: create stub sheet with banner only ──
        if not raw:
            df = pd.DataFrame([{
                "Date": "", "Invoice #": "—", "Type": "—",
                "Stmt Amount": None, "GL Amount": None, "Variance": None,
                "Status": "⚠ Extraction Failed",
            }])
            sheets[sn] = df
            sheet_meta[sn] = {
                "vendor_display":      vendor_display or gv or "Unknown",
                "entity_display":      entity_display or "Unknown",
                "stmt_date":           stmt_date or "",
                "source_file":         src,
                "vendor_stated_total": vendor_stated_total,
                "extracted_total":     0.0,
                "l1_variance":         None,
                "l1_status":           "⚠ Extraction Failed",
                "gl_total":            0.0,
                "net_variance":        0.0,
                "l2_status":           "— (no extraction)",
            }
            srows.append({
                "Statement":         sn,
                "Source File":       src,
                "Items":             0,
                "Matched":           0,
                "Amt Variance":      0,
                "Missing in GL":     0,
                "Vendor Stmt Total": round(vendor_stated_total, 2) if vendor_stated_total is not None else None,
                "Extracted Total":   0.0,
                "Layer 1 Variance":  None,
                "Layer 1 Status":    "⚠ Extraction Failed",
                "GL Total":          0.0,
                "Net Variance":      0.0,
                "Layer 2 Status":    "— (no extraction)",
            })
            reconciled.add(src)
            log(f"    ⚠ Extraction failed — stub sheet created for manual review")
            return

        # ── Build the data table: ALL rows appear (including payments) ──
        lk = glk(gl, gv, gl_l); recon = []
        for it in raw:
            inv = str(it["Invoice"]).strip()
            sa = round(it["Amount"], 2)
            type_str = it.get("Type", "Invoice")

            if type_str in SKIP_TYPES:
                # Payments / Unapplied Cash — show on sheet but skip GL lookup
                recon.append({
                    "Date": it["Date"], "Invoice #": inv, "Type": type_str,
                    "Stmt Amount": sa, "GL Amount": None, "Variance": None,
                    "Status": f"{type_str} — Not Reconciled",
                })
                continue

            mk, ga = mi(inv, lk)
            if mk is not None:
                ga = round(ga, 2); v = round(sa - ga, 2)
                st = "Matched" if abs(v) < 0.015 else "Amount Variance"
                recon.append({"Date": it["Date"], "Invoice #": inv, "Type": type_str,
                              "Stmt Amount": sa, "GL Amount": ga, "Variance": v, "Status": st})
            else:
                recon.append({"Date": it["Date"], "Invoice #": inv, "Type": type_str,
                              "Stmt Amount": sa, "GL Amount": None, "Variance": None,
                              "Status": "Missing in GL"})

        df = pd.DataFrame(recon); sheets[sn] = df

        # ── Layer 1: vendor stated total vs extracted total (sum of ALL rows) ──
        extracted_total = round(sum(r["Amount"] for r in raw), 2)
        if vendor_stated_total is not None:
            l1_var = round(extracted_total - vendor_stated_total, 2)
            if abs(l1_var) <= 0.02:
                l1_status = "✓ Reconciled"
            else:
                l1_status = "⚠ Manual Review Required"
                log(f"    ⚠ Layer 1 variance: extracted ${extracted_total:.2f} vs vendor ${vendor_stated_total:.2f} (diff ${l1_var:+.2f})")
        else:
            l1_var = None
            l1_status = "❓ Total Not Found"

        # ── Layer 2: GL comparison (count only invoice/credit-memo rows, not payments) ──
        recon_df = df[~df.Status.str.contains("Not Reconciled", na=False)]
        m_count = len(recon_df[recon_df.Status == "Matched"])
        v_count = len(recon_df[recon_df.Status == "Amount Variance"])
        mig_count = len(recon_df[recon_df.Status == "Missing in GL"])
        gl_total = float(recon_df["GL Amount"].sum()) if recon_df["GL Amount"].notna().any() else 0.0
        extracted_inv_total = float(recon_df["Stmt Amount"].sum())
        net_var = round(extracted_inv_total - gl_total, 2)

        if m_count + v_count + mig_count == 0:
            l2_status = "— (no invoices to reconcile)"
        elif v_count == 0 and mig_count == 0:
            l2_status = "✓ All Matched"
        else:
            issues = []
            if v_count > 0: issues.append(f"{v_count} variance{'s' if v_count > 1 else ''}")
            if mig_count > 0: issues.append(f"{mig_count} missing")
            l2_status = "⚠ " + ", ".join(issues)

        sheet_meta[sn] = {
            "vendor_display":       vendor_display or gv or "Unknown",
            "entity_display":       entity_display or "Unknown",
            "stmt_date":            stmt_date or "",
            "source_file":          src,
            "vendor_stated_total":  vendor_stated_total,
            "extracted_total":      extracted_total,
            "l1_variance":          l1_var,
            "l1_status":            l1_status,
            "gl_total":             round(gl_total, 2),
            "net_variance":         net_var,
            "l2_status":            l2_status,
        }

        srows.append({
            "Statement":         sn,
            "Source File":       src,
            "Items":             len(df),
            "Matched":           m_count,
            "Amt Variance":      v_count,
            "Missing in GL":     mig_count,
            "Vendor Stmt Total": round(vendor_stated_total, 2) if vendor_stated_total is not None else None,
            "Extracted Total":   round(extracted_total, 2),
            "Layer 1 Variance":  l1_var,
            "Layer 1 Status":    l1_status,
            "GL Total":          round(gl_total, 2),
            "Net Variance":      net_var,
            "Layer 2 Status":    l2_status,
        })
        reconciled.add(src)
        log(f"    {len(df)} items: {m_count} matched, {v_count} variance, {mig_count} missing | L1: {l1_status}")

    def process_rows(fn, fp, rows, fallback_label, fallback_vendor, fallback_locs,
                     stmt_total=None, stmt_date=None):
        """
        Route rows through smart-match → fallback. NEVER excludes a statement
        for Layer 1 failures. NEVER fabricates 'Payment Adjustment' entries.
        Layer 1 reconciliation happens inside do() and surfaces as a flag.
        """
        smart = smart_invoice_match(rows, gl, log)
        if smart and len(smart) == 1:
            # Single vendor/location identified — pass ALL rows (including payments
            # and Missing-in-GL invoices) so the sheet is complete and Layer 1 reconciles.
            sg = smart[0]
            do(sg["label"], rows, sg["vendor"], [sg["loc_id"]], fn,
               vendor_stated_total=stmt_total,
               vendor_display=sg["vendor"],
               entity_display=sg["loc_id"],
               stmt_date=stmt_date)
        elif smart:
            # Multi-vendor statement — split into per-vendor sheets, no per-sheet vendor total
            for sg in smart:
                do(sg["label"], sg["rows"], sg["vendor"], [sg["loc_id"]], fn,
                   vendor_stated_total=None,
                   vendor_display=sg["vendor"],
                   entity_display=sg["loc_id"],
                   stmt_date=stmt_date)
        else:
            log(f"    Smart match found nothing — using filename fallback")
            do(fallback_label, rows, fallback_vendor, fallback_locs, fn,
               vendor_stated_total=stmt_total,
               vendor_display=fallback_vendor,
               entity_display=fallback_label,
               stmt_date=stmt_date)

    # ──────────────────────────────────────────────────────────────────────
    #  Per-statement loop
    # ──────────────────────────────────────────────────────────────────────
    # Track which input files have produced at least one sheet.
    # HARD RULE: at the end, every input file must have ≥ 1 sheet — no exceptions.
    files_with_sheets = set()
    # Wrap do() so we can track which source files produced sheets
    _orig_do = do
    def do(sn, raw, gv, gl_l, src, **kwargs):
        result = _orig_do(sn, raw, gv, gl_l, src, **kwargs)
        files_with_sheets.add(src)
        return result

    def make_stub(fn, label_suffix, l1_status, vendor_stated=None,
                  vendor_display=None, entity_display=None):
        """Create a stub sheet for a file that couldn't be processed normally."""
        stub_label = (fn.replace(".pdf", "").replace(".xlsx", ""))[:31]
        do(stub_label, [], "", [], fn,
           vendor_stated_total=vendor_stated,
           vendor_display=vendor_display or "Unknown",
           entity_display=entity_display or "Unknown")

    for fn in sorted(stmts):
        try:
            fp   = stmt_map[fn]
            ov   = file_overrides.get(fn, {})
            l, v = fi(fn)
            log(f"Processing: {fn}")

            # ── US Paper: multi-entity statement with explicit per-entity sub-totals ──
            if v == "US PAPER":
                txt = _pdf(fp)
                us = parse_us_paper(txt)
                section_totals = find_section_totals(txt)
                sub_total_map = {}
                for s in section_totals:
                    sub_total_map[s["section"].strip().upper()] = s["total"]

                def lookup_subtotal(canonical_name):
                    up = canonical_name.upper()
                    if up in sub_total_map:
                        return sub_total_map[up]
                    for ku, t in sub_total_map.items():
                        if up.startswith(ku) and (len(ku) >= 8 or len(ku.split()) >= 2):
                            return t
                    for ku, t in sub_total_map.items():
                        if ku.startswith(up) and (len(up) >= 5 or len(up.split()) >= 2):
                            return t
                    return None

                usm = {
                    "MAD DOGS AND ENGLISHMEN": (["MAD-80041"], "MD Us Paper", "MAD DOGS AND ENGLISHMEN"),
                    "OXFORD EXCHANGE LLC":     (["OE","OE-96001","OE-96003","OE-96004","OE-96005","OE-96008","OE-96011"], "OE Us Paper", "OXFORD EXCHANGE LLC"),
                    "Predalina LLC":           (["PRED","PRED-82000"], "PRED Us Paper", "Predalina LLC"),
                    "SH-19":                   (["SH-93004"], "SH19 Us Paper", "SH-19"),
                    "The Library St Pete":     (["LIB-96100"], "LIB Us Paper", "The Library St Pete"),
                    "The Stovall House":       (["SH-93001","SH-93002"], "SH Us Paper", "The Stovall House"),
                }
                grand_total, _ = find_statement_total(txt)
                found_sub = False
                for canonical, ir in us.items():
                    if canonical in usm and ir:
                        locs2, sn2, entity_disp = usm[canonical]
                        sub_total = lookup_subtotal(canonical)
                        log(f"  Sub: {canonical} (vendor sub-total: ${sub_total:.2f})" if sub_total else f"  Sub: {canonical}")
                        do(sn2, ir, "US PAPER CORP", locs2, fn,
                           vendor_stated_total=sub_total,
                           vendor_display="US PAPER CORP",
                           entity_display=entity_disp)
                        found_sub = True
                if not found_sub:
                    log(f"  ⚠ No US Paper entity sections matched — creating stub")
                    make_stub(fn, "US Paper",
                              l1_status="⚠ Extraction Failed",
                              vendor_stated=grand_total,
                              vendor_display="US PAPER CORP",
                              entity_display=l or "Unknown")
                continue

            if fn.endswith(".xlsx") and "AMAZON" in fn.upper():
                r = parse_amazon_xl(fp)
                gv   = ov.get("gl_vendor") or VM.get(v, "Amazon Capital Services")
                gl_l = ov.get("gl_locs")   or LOC.get(l, [])
                if r:
                    do(f"{l or ''} Amazon".strip(), r, gv, gl_l, fn,
                       vendor_display=gv, entity_display=l or "")
                else:
                    log(f"  ⚠ Amazon XLSX parse returned no rows — creating stub")
                    make_stub(fn, "Amazon",
                              l1_status="⚠ Extraction Failed",
                              vendor_display="Amazon Capital Services",
                              entity_display=l or "")
                continue

            txt = _pdf(fp)

            # STEP 1: Find vendor's authoritative statement total
            stmt_total, total_label = find_statement_total(txt)
            if stmt_total:
                log(f"    Vendor stated total: ${stmt_total:,.2f} (via {total_label})")
            else:
                log(f"    ❓ Vendor total not found in statement")

            # STEP 2: Extract invoices
            parser = PARSERS.get(v) if v else None
            rows = None
            if parser and parser != parse_us_paper:
                rows = parser(txt)
            if not rows:
                rows = parse_wrights(txt)
            if not rows:
                rows = parse_generic(txt)

            if not rows and not txt.strip():
                log(f"  ⚠ No text extracted from PDF — creating stub sheet")
                make_stub(fn, "PDF unreadable",
                          l1_status="⚠ Extraction Failed",
                          vendor_display="Unknown (PDF unreadable)",
                          entity_display=l or "Unknown")
                continue
            if not rows:
                log(f"  ⚠ No invoices extracted ({len(txt)} chars of text) — creating stub sheet")
                make_stub(fn, "extraction failed",
                          l1_status="⚠ Extraction Failed",
                          vendor_stated=stmt_total,
                          vendor_display="Unknown (extraction failed)",
                          entity_display=l or "Unknown")
                continue

            fb_vendor = ov.get("gl_vendor") or (VM.get(v) if v else "") or ""
            if isinstance(fb_vendor, list): fb_vendor = fb_vendor[0]
            fb_locs   = ov.get("gl_locs")   or (LOC.get(l, []) if l else [])
            fb_label  = f"{l} {v.title()}" if l and v else fn.replace(".pdf", "").replace(".xlsx", "")[:31]

            process_rows(fn, fp, rows, fb_label, fb_vendor, fb_locs,
                         stmt_total=stmt_total)
        except Exception as e:
            log(f"  ❌ EXCEPTION processing {fn}: {type(e).__name__}: {e}")
            try:
                make_stub(fn, "exception",
                          l1_status="❌ Error",
                          vendor_display=f"Error: {type(e).__name__}",
                          entity_display="See logs")
            except Exception as e2:
                log(f"  ❌ Could not even create stub: {e2}")

    # ── HARD GUARD: every input file must have at least one sheet ──
    for fn in sorted(stmts):
        if fn not in files_with_sheets:
            log(f"  ⚠ Final guard: {fn} produced no sheet — creating stub")
            try:
                make_stub(fn, "guard",
                          l1_status="⚠ No sheet produced",
                          vendor_display="Unknown",
                          entity_display="Unknown")
            except Exception as e:
                log(f"  ❌ Final guard could not create stub: {e}")

    if not srows:
        raise ValueError("No vendor statements were successfully processed.")

    _today = date.today().strftime("%m%d%y")
    output_filename = f"AP_RECONCILIATION_{_today}.xlsx"

    log(f"Building workbook…")
    sdf = pd.DataFrame(srows)
    BANNER_ROWS = 13  # height of the metadata banner above each detail table

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        sdf.to_excel(w, sheet_name="Summary", index=False)
        for sn in sorted(sheets):
            sheets[sn].to_excel(w, sheet_name=sn[:31], index=False, startrow=BANNER_ROWS)
    buf.seek(0)

    wb = load_workbook(buf)

    # Write banner content + format each detail sheet
    for sn, df in sheets.items():
        s = sn[:31]
        if s in wb.sheetnames:
            ws = wb[s]
            meta = sheet_meta.get(sn, {})
            _write_banner(ws, meta, BANNER_ROWS)
            fmt_detail(ws, len(df), len(df.columns), banner_rows=BANNER_ROWS)

    ws = wb["Summary"]
    fmt_summary(ws, len(sdf), len(sdf.columns))

    # Hyperlinks from Summary → detail sheets
    for r in range(2, len(sdf) + 2):
        cl = ws.cell(r, 1); sn = cl.value
        if sn and sn[:31] in wb.sheetnames:
            cl.hyperlink = f"#'{sn[:31]}'!A1"; cl.font = _AL

    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)

    tm = sdf["Matched"].sum(); tv = sdf["Amt Variance"].sum(); tmi = sdf["Missing in GL"].sum()
    l1_ok    = (sdf["Layer 1 Status"] == "✓ Reconciled").sum()
    l1_warn  = (sdf["Layer 1 Status"] == "⚠ Manual Review Required").sum()
    l1_miss  = (sdf["Layer 1 Status"] == "❓ Total Not Found").sum()
    log(f"Done! {len(sheets)+1} sheets — Layer 1: {l1_ok} reconciled / {l1_warn} review / {l1_miss} no-total | Layer 2: {tm} matched | {tv} variances | {tmi} missing")

    skipped = sorted(all_stmt_set - reconciled)
    return out_buf.getvalue(), output_filename, reconciled, skipped
