"""
Expense Trends — distributable Excel report for operations.
Path: backend/trends_report.py

Builds a styled workbook from the same analysis that powers the web page:
  Sheet 1 "Expense Trends" — vendor x month grid, sectioned by GL group in
    P&L order. Orange header band, dark group bands with per-column SUM
    formulas, flagged vendor rows tinted (amber = possibly high, red =
    possibly missing/low), analysis-month column emphasized, spillover
    (+1) month included, TOTAL column per row, grand-total row.
  Sheet 2 "Review Queue" — every flag, worst first, with its statistics.

All totals are live Excel formulas (SUM), never hardcoded, so the file
stays auditable and recalculates if ops filters or edits it.
"""
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from trends_engine import analyze

# Caspers palette (matches the AP Rec output styling)
OX_ORANGE = "FF7030"
DARK_BAND = "1E1B17"
EMPH_FILL = "FFE8DC"   # analysis-month column tint
AMBER     = "FFF3CD"   # possibly high
RED       = "FCE8E6"   # possibly missing / low
WHITE     = "FFFFFF"

MONEY_FMT = '$#,##0;($#,##0);"-"'
def F(**kw):
    kw.setdefault("name", "Arial")
    kw.setdefault("size", 10)
    return Font(**kw)
FILL = lambda c: PatternFill("solid", start_color=c)
THIN = Side(style="thin", color="D9D2C8")
BORDER = Border(bottom=THIN)


def _mlabel(m):
    y, mo = m.split("-")
    return f"{['','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'][int(mo)]} '{y[2:]}"


def build_report(gl_path, entity=None, view="vendor", period=None) -> tuple[bytes, str]:
    data = analyze(gl_path, entity=entity, view=view, period=period)
    months = data["months"]
    n_m = len(months)
    analysis_idx = months.index(data["period"])
    tot_col = n_m + 2                 # A=label, B.. months, then TOTAL
    flag_col = tot_col + 1
    last_letter = get_column_letter(flag_col)

    wb = Workbook()

    # ── Sheet 1: the grid ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Expense Trends"
    ent_label = (entity or "ALL ENTITIES (ORG)").upper()
    ws["A1"] = f"EXPENSE TRENDS — {ent_label}"
    ws["A1"].font = F(bold=True, size=14)
    ws["A2"] = (f"Analysis month: {_mlabel(data['period'])}   ·   window "
                f"{_mlabel(months[0])} – {_mlabel(months[-1])} (last column = next-month "
                f"spillover)   ·   generated {date.today().isoformat()}   ·   "
                f"view: by {view}")
    ws["A2"].font = F(size=9, color="8C7B6A")

    # Header band
    hr = 4
    ws.cell(row=hr, column=1, value="VENDOR" if view == "vendor" else "ACCOUNT")
    for i, m in enumerate(months):
        ws.cell(row=hr, column=2 + i, value=_mlabel(m))
    ws.cell(row=hr, column=tot_col, value="TOTAL")
    ws.cell(row=hr, column=flag_col, value="FLAG")
    for c in range(1, flag_col + 1):
        cell = ws.cell(row=hr, column=c)
        cell.font = F(bold=True, color=WHITE)
        cell.fill = FILL(OX_ORANGE)
        cell.alignment = Alignment(horizontal="right" if 1 < c < flag_col else "left")
    ws.cell(row=hr, column=2 + analysis_idx).font = F(bold=True, color=WHITE, underline="single")

    flag_fill = {"Possibly High": AMBER, "Possibly Missing": RED, "Possibly Low": RED}

    r = hr + 1
    group_band_rows = []
    for g in data["groups"]:
        band = r
        group_band_rows.append(band)
        first_vendor = band + 1
        ws.cell(row=band, column=1, value=g["name"])
        for _, row in enumerate(g["rows"]):
            r_v = first_vendor + _
            ws.cell(row=r_v, column=1, value=row["label"]).font = F()
            for i, v in enumerate(row["values"]):
                c = ws.cell(row=r_v, column=2 + i, value=v if v else None)
                c.number_format = MONEY_FMT
                c.font = F(bold=(i == analysis_idx))
            tc = ws.cell(row=r_v, column=tot_col,
                         value=f"=SUM(B{r_v}:{get_column_letter(1 + n_m)}{r_v})")
            tc.number_format = MONEY_FMT
            tc.font = F(bold=True)
            fl = row.get("flag")
            row_tint = flag_fill.get(fl["flag"]) if fl else None
            if fl:
                ws.cell(row=r_v, column=flag_col, value=fl["flag"]).font = F(size=9)
            if row_tint:
                for c in range(1, flag_col + 1):
                    ws.cell(row=r_v, column=c).fill = FILL(row_tint)
            else:
                ws.cell(row=r_v, column=2 + analysis_idx).fill = FILL(EMPH_FILL)
        last_vendor = first_vendor + len(g["rows"]) - 1
        # group band: dark, with live SUM formulas over its vendor rows
        for c in range(1, flag_col + 1):
            cell = ws.cell(row=band, column=c)
            cell.fill = FILL(DARK_BAND)
            cell.font = F(bold=True, color=WHITE)
        for i in range(n_m):
            col = get_column_letter(2 + i)
            cell = ws.cell(row=band, column=2 + i,
                           value=f"=SUM({col}{first_vendor}:{col}{last_vendor})")
            cell.number_format = MONEY_FMT
            cell.font = F(bold=True, color=WHITE)
        tcol = get_column_letter(tot_col)
        cell = ws.cell(row=band, column=tot_col,
                       value=f"=SUM({tcol}{first_vendor}:{tcol}{last_vendor})")
        cell.number_format = MONEY_FMT
        cell.font = F(bold=True, color=WHITE)
        r = last_vendor + 1

    # Grand total = sum of the group band rows only (no double counting)
    gr = r + 1
    ws.cell(row=gr, column=1, value="GRAND TOTAL").font = F(bold=True, size=11)
    for i in range(n_m):
        col = get_column_letter(2 + i)
        cell = ws.cell(row=gr, column=2 + i,
                       value="=" + "+".join(f"{col}{b}" for b in group_band_rows))
        cell.number_format = MONEY_FMT
        cell.font = F(bold=True, size=11)
    tcol = get_column_letter(tot_col)
    cell = ws.cell(row=gr, column=tot_col,
                   value="=" + "+".join(f"{tcol}{b}" for b in group_band_rows))
    cell.number_format = MONEY_FMT
    cell.font = F(bold=True, size=11)
    for c in range(1, flag_col + 1):
        ws.cell(row=gr, column=c).border = Border(top=Side(style="double", color=DARK_BAND))

    ws.column_dimensions["A"].width = 38
    for i in range(n_m):
        ws.column_dimensions[get_column_letter(2 + i)].width = 11
    ws.column_dimensions[get_column_letter(tot_col)].width = 13
    ws.column_dimensions[get_column_letter(flag_col)].width = 17
    ws.freeze_panes = f"B{hr + 1}"
    ws.sheet_view.showGridLines = False

    # ── Sheet 2: Review Queue ──────────────────────────────────────────────
    q = wb.create_sheet("Review Queue")
    q["A1"] = f"REVIEW QUEUE — {len(data['flags'])} flags, worst first"
    q["A1"].font = F(bold=True, size=13)
    headers = ["FLAG", "GL GROUP", "VENDOR" if view == "vendor" else "ACCOUNT",
               "CURRENT", "HISTORY AVG", "STD DEV", "MONTHS PRESENT", "LAST DOC #"]
    for c, h in enumerate(headers, 1):
        cell = q.cell(row=3, column=c, value=h)
        cell.font = F(bold=True, color=WHITE)
        cell.fill = FILL(OX_ORANGE)
    for i, fl in enumerate(data["flags"]):
        rr = 4 + i
        vals = [fl["flag"], fl["group"], fl["label"], fl["current"],
                fl["history_mean"], fl["history_sd"],
                f"{fl['months_present']}/{fl['months_history']}", fl.get("last_doc", "")]
        for c, v in enumerate(vals, 1):
            cell = q.cell(row=rr, column=c, value=v)
            cell.font = F(size=9)
            if c in (4, 5, 6):
                cell.number_format = MONEY_FMT
        tint = flag_fill.get(fl["flag"])
        if tint:
            for c in range(1, len(headers) + 1):
                q.cell(row=rr, column=c).fill = FILL(tint)
    for c, w in zip("ABCDEFGH", (17, 24, 36, 13, 13, 12, 15, 22)):
        q.column_dimensions[c].width = w
    q.freeze_panes = "A4"
    q.sheet_view.showGridLines = False

    buf = BytesIO()
    wb.save(buf)
    fname = f"EXPENSE_TRENDS_{(entity or 'ORG').upper()}_{data['period']}.xlsx"
    return buf.getvalue(), fname
